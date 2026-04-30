from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


MONTHS: dict[str, int] = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


class BRInfoConfig(BaseModel):
    """Agent-adjustable extraction rules for accountant BR override workbooks."""

    sheet_name: str | None = None
    sheet_name_keywords: list[str] = Field(default_factory=lambda: [])
    label_column: int = 1
    max_header_scan_rows: int = 12
    minimum_month_headers: int = 3
    include_blank_values: bool = False
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "BRInfoConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class BRInfoExtraction:
    path: Path
    sheet_name: str
    header_row: int
    year: int | None
    rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    @property
    def overrides(self) -> pl.DataFrame:
        return pl.DataFrame(self.rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "year": self.year,
            "warnings": self.warnings,
            "rows": self.rows,
        }


def extract_br_info(path: Path, config: BRInfoConfig | None = None) -> BRInfoExtraction:
    config = config or BRInfoConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        sheet_name = choose_sheet(workbook_values.sheetnames, config)
        if sheet_name is None:
            raise ValueError("Could not find BR Info sheet")

        ws_values = workbook_values[sheet_name]
        ws_formulas = workbook_formulas[sheet_name]
        header_row, month_columns = detect_month_header_row(ws_values, config)
        year = detect_year(ws_values, sheet_name, header_row)
        rows = extract_override_rows(
            path=path,
            ws_values=ws_values,
            ws_formulas=ws_formulas,
            header_row=header_row,
            year=year,
            month_columns=month_columns,
            config=config,
            formula_evaluator=formula_evaluator,
        )
        warnings = validate_extraction(rows, month_columns)
        return BRInfoExtraction(
            path=path,
            sheet_name=sheet_name,
            header_row=header_row,
            year=year,
            rows=rows,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def choose_sheet(sheet_names: list[str], config: BRInfoConfig) -> str | None:
    if config.sheet_name:
        return config.sheet_name if config.sheet_name in sheet_names else None
    if not config.sheet_name_keywords:
        return sheet_names[0] if sheet_names else None

    scored: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        normalized = normalize_key(sheet_name)
        score = sum(100 for keyword in config.sheet_name_keywords if normalize_key(keyword) in normalized)
        scored.append((score, sheet_name))
    scored.sort(reverse=True)
    return scored[0][1] if scored and scored[0][0] else None


def detect_month_header_row(
    ws: Worksheet,
    config: BRInfoConfig,
) -> tuple[int, dict[int, tuple[str, int]]]:
    candidates: list[tuple[int, int, dict[int, tuple[str, int]]]] = []
    for row_idx in range(1, min(ws.max_row, config.max_header_scan_rows) + 1):
        month_columns: dict[int, tuple[str, int]] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = normalize_text(ws.cell(row_idx, col_idx).value)
            month_num = month_number(value)
            if month_num is not None:
                month_columns[col_idx] = (month_name(month_num), month_num)
        if len(month_columns) >= config.minimum_month_headers:
            candidates.append((len(month_columns), row_idx, month_columns))

    if not candidates:
        raise ValueError("Could not detect month header row")
    candidates.sort(reverse=True)
    _, row_idx, month_columns = candidates[0]
    return row_idx, month_columns


def detect_year(ws: Worksheet, sheet_name: str, header_row: int) -> int | None:
    sheet_year = year_from_value(sheet_name)
    if sheet_year is not None:
        return sheet_year
    for row_idx in range(1, header_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            year = year_from_value(ws.cell(row_idx, col_idx).value)
            if year is not None:
                return year
    return None


def extract_override_rows(
    path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    header_row: int,
    year: int | None,
    month_columns: dict[int, tuple[str, int]],
    config: BRInfoConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        override_name = normalize_text(ws_values.cell(row_idx, config.label_column).value)
        if not override_name:
            continue

        for col_idx, (month, month_num) in sorted(month_columns.items()):
            value_info = formula_cell_info(
                ws_values,
                ws_formulas,
                row_idx,
                col_idx,
                formula_evaluator,
                config,
            )
            value = value_info["amount"]
            if value in (None, "") and not config.include_blank_values:
                continue
            rows.append(
                {
                    "source_file": str(path),
                    "sheet": ws_values.title,
                    "year": year,
                    "month_num": month_num,
                    "month_name": month,
                    "override_name": override_name,
                    "value": to_float(value),
                    "raw_value": normalize_text(value),
                    "source_cell": f"{get_column_letter(col_idx)}{row_idx}",
                    "formula": value_info["formula"],
                    "cached_value": to_float(value_info["cached_value"]),
                    "calculated_value": to_float(value_info["calculated_value"]),
                    "calculation_status": value_info["calculation_status"],
                    "calculation_detail": value_info["calculation_detail"],
                }
            )
    return rows


def formula_cell_info(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    col_idx: int,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    config: BRInfoConfig,
) -> dict[str, Any]:
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    calculated_value = None
    status = "not_formula"
    detail = None
    amount = cached_value

    if formula_evaluator is not None and formula_cell.data_type == "f":
        calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
        calculated_value = calculation.value
        status = calculation.status
        detail = calculation.detail
        if (
            config.use_calculated_formula_values
            and calculation.status == "ok"
            and not is_formula_sentinel(calculation.value)
        ):
            amount = calculation.value

    return {
        "amount": amount,
        "formula": formula,
        "cached_value": cached_value,
        "calculated_value": calculated_value,
        "calculation_status": status,
        "calculation_detail": detail,
    }


def validate_extraction(rows: list[dict[str, Any]], month_columns: dict[int, tuple[str, int]]) -> list[str]:
    warnings: list[str] = []
    if not month_columns:
        warnings.append("No month columns detected")
    if not rows:
        warnings.append("No BR override rows parsed")
    return warnings


def month_number(value: str) -> int | None:
    normalized = normalize_key(value)
    if normalized in MONTHS:
        return MONTHS[normalized]
    return None


def month_name(month_num: int) -> str:
    for name, number in MONTHS.items():
        if number == month_num:
            return name.capitalize()
    raise ValueError(f"Invalid month number: {month_num}")


def year_from_value(value: Any) -> int | None:
    if isinstance(value, int) and 1900 <= value <= 2200:
        return value
    text = normalize_text(value)
    match = re.search(r"\b(20\d{2}|19\d{2})\b", text)
    if match:
        return int(match.group(1))
    return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(BRInfoConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
