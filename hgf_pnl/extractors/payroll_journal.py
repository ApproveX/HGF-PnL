from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


class PayrollJournalConfig(BaseModel):
    """Agent-adjustable extraction rules for payroll journal workbooks."""

    payroll_sheet_keywords: list[str] = Field(default_factory=lambda: ["payroll"])
    distribution_sheet_keywords: list[str] = Field(default_factory=lambda: ["distribution"])
    employee_code_column: int = 1
    employee_name_column: int = 2
    gross_pay_column: int = 3
    section_label_column: int = 4
    section_total_column: int = 5
    allocation_min_column: int = 7
    allocation_max_column: int = 14
    allocation_total_patterns: list[str] = Field(default_factory=lambda: [r"^total$"])
    distribution_amount_column: int = 2
    stop_distribution_after_blank_rows: int = 4
    preserve_zero_allocations: bool = False
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True
    derive_distribution_from_payroll_sheet: bool = True
    use_calculated_distribution_formula_values: bool = False

    @classmethod
    def from_json_file(cls, path: Path | None) -> "PayrollJournalConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class PayrollJournalExtraction:
    path: Path
    payroll_sheet: str
    distribution_sheet: str | None
    employee_rows: list[dict[str, Any]]
    allocation_rows: list[dict[str, Any]]
    allocation_summary_rows: list[dict[str, Any]]
    distribution_rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    @property
    def employees(self) -> pl.DataFrame:
        return pl.DataFrame(self.employee_rows)

    @property
    def allocations(self) -> pl.DataFrame:
        return pl.DataFrame(self.allocation_rows)

    @property
    def allocation_summaries(self) -> pl.DataFrame:
        return pl.DataFrame(self.allocation_summary_rows)

    @property
    def distribution(self) -> pl.DataFrame:
        return pl.DataFrame(self.distribution_rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "payroll_sheet": self.payroll_sheet,
            "distribution_sheet": self.distribution_sheet,
            "warnings": self.warnings,
            "employee_rows": self.employee_rows,
            "allocation_rows": self.allocation_rows,
            "allocation_summary_rows": self.allocation_summary_rows,
            "distribution_rows": self.distribution_rows,
        }


def extract_payroll_journal(
    path: Path,
    config: PayrollJournalConfig | None = None,
) -> PayrollJournalExtraction:
    config = config or PayrollJournalConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        warnings: list[str] = []
        payroll_sheet = choose_sheet(workbook_values.sheetnames, config.payroll_sheet_keywords, prefer_exact="payroll")
        if payroll_sheet is None:
            raise ValueError("Could not find payroll sheet")

        distribution_sheet = choose_sheet(
            workbook_values.sheetnames,
            config.distribution_sheet_keywords,
            prefer_exact="payroll distribution",
        )
        if distribution_sheet is None:
            warnings.append("Could not find payroll distribution sheet")

        ws_values = workbook_values[payroll_sheet]
        ws_formulas = workbook_formulas[payroll_sheet]
        section_by_row, section_totals = detect_employee_sections(ws_values, ws_formulas, config, formula_evaluator)
        allocation_context = extract_allocation_rows(
            path,
            ws_values,
            ws_formulas,
            config,
            section_by_row,
            formula_evaluator,
        )
        employee_rows = extract_employee_rows(
            path,
            ws_values,
            ws_formulas,
            config,
            section_by_row,
            section_totals,
            allocation_context.allocated_total_by_row,
            formula_evaluator,
        )

        distribution_rows: list[dict[str, Any]] = []
        if config.derive_distribution_from_payroll_sheet:
            distribution_rows = extract_distribution_rows_from_payroll_sheet(
                path,
                ws_values,
                ws_formulas,
                config,
                formula_evaluator,
            )
        elif distribution_sheet is not None:
            distribution_rows = extract_distribution_rows(
                path,
                workbook_values[distribution_sheet],
                workbook_formulas[distribution_sheet],
                config,
                formula_evaluator,
            )

        warnings.extend(validate_extraction(employee_rows, allocation_context.rows, distribution_rows))
        return PayrollJournalExtraction(
            path=path,
            payroll_sheet=payroll_sheet,
            distribution_sheet=distribution_sheet,
            employee_rows=employee_rows,
            allocation_rows=allocation_context.rows,
            allocation_summary_rows=allocation_context.summary_rows,
            distribution_rows=distribution_rows,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


@dataclass
class AllocationExtraction:
    rows: list[dict[str, Any]]
    summary_rows: list[dict[str, Any]]
    allocated_total_by_row: dict[int, float]


def choose_sheet(sheet_names: list[str], keywords: list[str], prefer_exact: str | None = None) -> str | None:
    normalized_preference = normalize_key(prefer_exact or "")
    scored: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        normalized = normalize_key(sheet_name)
        score = 0
        if normalized_preference and normalized == normalized_preference:
            score += 1000
        score += sum(100 for keyword in keywords if normalize_key(keyword) in normalized)
        scored.append((score, sheet_name))
    scored.sort(reverse=True)
    if not scored or scored[0][0] == 0:
        return None
    return scored[0][1]


def detect_employee_sections(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> tuple[dict[int, str | None], dict[str, dict[str, Any]]]:
    section_by_row: dict[int, str | None] = {}
    section_totals: dict[str, dict[str, Any]] = {}
    pending_rows: list[int] = []

    for row_idx in range(1, ws_values.max_row + 1):
        if not is_employee_row(ws_values, row_idx, config):
            continue
        pending_rows.append(row_idx)
        section = normalize_text(ws_values.cell(row_idx, config.section_label_column).value) or None
        if section is None:
            continue

        total_info = formula_cell_info(
            ws_values,
            ws_formulas,
            row_idx,
            config.section_total_column,
            formula_evaluator,
            config,
        )
        section_totals[section] = {
            "source_row": row_idx,
            "amount": to_float(total_info["amount"]),
            "formula": total_info["formula"],
            "cached_amount": to_float(total_info["cached_amount"]),
            "calculated_amount": to_float(total_info["calculated_amount"]),
            "calculation_status": total_info["calculation_status"],
            "calculation_detail": total_info["calculation_detail"],
        }
        for employee_row in pending_rows:
            section_by_row[employee_row] = section
        pending_rows = []

    for employee_row in pending_rows:
        section_by_row[employee_row] = None
    return section_by_row, section_totals


def extract_employee_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    section_by_row: dict[int, str | None],
    section_totals: dict[str, dict[str, Any]],
    allocated_total_by_row: dict[int, float],
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(1, ws_values.max_row + 1):
        if not is_employee_row(ws_values, row_idx, config):
            continue
        gross_info = formula_cell_info(
            ws_values,
            ws_formulas,
            row_idx,
            config.gross_pay_column,
            formula_evaluator,
            config,
        )
        gross_pay = to_float(gross_info["amount"])
        allocated_total = allocated_total_by_row.get(row_idx)
        section = section_by_row.get(row_idx)
        section_total = section_totals.get(section or "", {})
        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "row": row_idx,
                "employee_code": normalize_text(ws_values.cell(row_idx, config.employee_code_column).value),
                "employee_name": normalize_text(ws_values.cell(row_idx, config.employee_name_column).value),
                "gross_pay": gross_pay,
                "section": section,
                "department": section,
                "department_source_row": section_total.get("source_row"),
                "section_total_amount": section_total.get("amount"),
                "section_total_source_row": section_total.get("source_row"),
                "department_total_amount": section_total.get("amount"),
                "allocated_total": allocated_total,
                "allocation_difference": gross_pay - allocated_total
                if gross_pay is not None and allocated_total is not None
                else None,
                "formula": gross_info["formula"],
                "cached_gross_pay": to_float(gross_info["cached_amount"]),
                "calculated_gross_pay": to_float(gross_info["calculated_amount"]),
                "calculation_status": gross_info["calculation_status"],
                "calculation_detail": gross_info["calculation_detail"],
            }
        )
    return rows


def extract_allocation_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    section_by_row: dict[int, str | None],
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> AllocationExtraction:
    rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    allocated_total_by_row: dict[int, float] = {}
    for header_row in find_allocation_header_rows(ws_values, config):
        headers = allocation_headers(ws_values, header_row, config)
        total_columns = [
            col_idx
            for col_idx, label in headers.items()
            if matches_any(label, config.allocation_total_patterns)
        ]
        total_col = max(total_columns) if total_columns else None

        employee_row_indices: list[int] = []
        row_idx = header_row + 1
        while row_idx <= ws_values.max_row and is_employee_row(ws_values, row_idx, config):
            employee_row_indices.append(row_idx)
            employee_name = normalize_text(ws_values.cell(row_idx, config.employee_name_column).value)
            gross_pay = to_float(ws_values.cell(row_idx, config.gross_pay_column).value)
            department = section_by_row.get(row_idx)

            if total_col is not None:
                total_info = formula_cell_info(
                    ws_values,
                    ws_formulas,
                    row_idx,
                    total_col,
                    formula_evaluator,
                    config,
                )
                total_value = to_float(total_info["amount"])
                if total_value is not None:
                    allocated_total_by_row[row_idx] = total_value

            for col_idx, label in headers.items():
                if matches_any(label, config.allocation_total_patterns):
                    continue
                value_info = formula_cell_info(
                    ws_values,
                    ws_formulas,
                    row_idx,
                    col_idx,
                    formula_evaluator,
                    config,
                )
                amount = to_float(value_info["amount"])
                if amount in (None, 0.0) and not config.preserve_zero_allocations:
                    continue
                rows.append(
                    {
                        "source_file": str(source_path),
                        "sheet": ws_values.title,
                        "header_row": header_row,
                        "row": row_idx,
                        "employee_code": normalize_text(ws_values.cell(row_idx, config.employee_code_column).value),
                        "employee_name": employee_name,
                        "section": department,
                        "department": department,
                        "allocation_category": label,
                        "amount": amount,
                        "gross_pay": gross_pay,
                        "percent_of_gross": amount / gross_pay
                        if amount is not None and gross_pay not in (None, 0)
                        else None,
                        "formula": value_info["formula"],
                        "cached_amount": to_float(value_info["cached_amount"]),
                        "calculated_amount": to_float(value_info["calculated_amount"]),
                        "calculation_status": value_info["calculation_status"],
                        "calculation_detail": value_info["calculation_detail"],
                    }
                )
            row_idx += 1

        summary_rows.extend(
            extract_allocation_summary_rows(
                source_path,
                ws_values,
                ws_formulas,
                config,
                formula_evaluator,
                header_row,
                row_idx,
                headers,
                total_columns,
                employee_row_indices,
                section_by_row,
            )
        )
    return AllocationExtraction(rows=rows, summary_rows=summary_rows, allocated_total_by_row=allocated_total_by_row)


def extract_allocation_summary_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    header_row: int,
    total_row: int,
    headers: dict[int, str],
    total_columns: list[int],
    employee_row_indices: list[int],
    section_by_row: dict[int, str | None],
) -> list[dict[str, Any]]:
    if not employee_row_indices or total_row > ws_values.max_row:
        return []

    department = infer_allocation_block_department(employee_row_indices, section_by_row)
    if not department:
        return []

    rows: list[dict[str, Any]] = []
    for col_idx, label in headers.items():
        value_info = formula_cell_info(
            ws_values,
            ws_formulas,
            total_row,
            col_idx,
            formula_evaluator,
            config,
        )
        amount = to_float(value_info["amount"])
        if amount in (None, 0.0) and not config.preserve_zero_allocations:
            continue
        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "header_row": header_row,
                "total_row": total_row,
                "employee_start_row": min(employee_row_indices),
                "employee_end_row": max(employee_row_indices),
                "section": department,
                "department": department,
                "allocation_category": label,
                "amount": amount,
                "is_total_category": col_idx in total_columns or matches_any(label, config.allocation_total_patterns),
                "formula": value_info["formula"],
                "cached_amount": to_float(value_info["cached_amount"]),
                "calculated_amount": to_float(value_info["calculated_amount"]),
                "calculation_status": value_info["calculation_status"],
                "calculation_detail": value_info["calculation_detail"],
                "source_cell": ws_values.cell(total_row, col_idx).coordinate,
                "source_kind": "allocation_total_row",
            }
        )
    return rows


def infer_allocation_block_department(
    employee_row_indices: list[int],
    section_by_row: dict[int, str | None],
) -> str | None:
    sections = [section_by_row.get(row_idx) for row_idx in employee_row_indices]
    non_blank = [section for section in sections if section]
    if not non_blank:
        return None
    counts: dict[str, int] = {}
    for section in non_blank:
        counts[section] = counts.get(section, 0) + 1
    return sorted(counts.items(), key=lambda item: (item[1], item[0]), reverse=True)[0][0]


def find_allocation_header_rows(ws: Worksheet, config: PayrollJournalConfig) -> list[int]:
    header_rows: list[int] = []
    for row_idx in range(1, ws.max_row + 1):
        headers = allocation_headers(ws, row_idx, config)
        if len(headers) < 2:
            continue
        if not any(matches_any(label, config.allocation_total_patterns) for label in headers.values()):
            continue
        if normalize_text(ws.cell(row_idx, config.employee_name_column).value):
            continue
        header_rows.append(row_idx)
    return header_rows


def allocation_headers(ws: Worksheet, row_idx: int, config: PayrollJournalConfig) -> dict[int, str]:
    return {
        col_idx: normalize_text(ws.cell(row_idx, col_idx).value)
        for col_idx in range(config.allocation_min_column, min(config.allocation_max_column, ws.max_column) + 1)
        if normalize_text(ws.cell(row_idx, col_idx).value)
    }


def extract_distribution_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_block: str | None = None
    blank_run = 0
    previous_label: str | None = None

    for row_idx in range(1, ws_values.max_row + 1):
        label_text = normalize_text(ws_values.cell(row_idx, 1).value)
        amount_cell = ws_values.cell(row_idx, config.distribution_amount_column)
        amount_text = normalize_text(amount_cell.value)

        if label_text and not amount_text:
            current_block = label_text
            previous_label = None
            blank_run = 0
            continue

        if not label_text and not amount_text:
            blank_run += 1
            if blank_run >= config.stop_distribution_after_blank_rows:
                current_block = None
                previous_label = None
            continue
        blank_run = 0

        if current_block is None:
            continue

        value_info = formula_cell_info(
            ws_values,
            ws_formulas,
            row_idx,
            config.distribution_amount_column,
            formula_evaluator,
            config,
            use_calculated_formula_values=config.use_calculated_distribution_formula_values,
        )
        label = label_text or inferred_blank_distribution_label(previous_label, value_info["formula"])
        amount = to_float(value_info["amount"])
        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "block": current_block,
                "row": row_idx,
                "label": label,
                "amount": amount,
                "is_total_row": bool(re.search(r"\btotal\b", label, flags=re.IGNORECASE)),
                "is_check_row": bool(re.search(r"\bcheck\b", label, flags=re.IGNORECASE)),
                "is_difference_row": label.lower() == "difference",
                "formula": value_info["formula"],
                "cached_amount": to_float(value_info["cached_amount"]),
                "calculated_amount": to_float(value_info["calculated_amount"]),
                "calculation_status": value_info["calculation_status"],
                "calculation_detail": value_info["calculation_detail"],
            }
        )
        previous_label = label
    return rows


def extract_distribution_rows_from_payroll_sheet(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rows.extend(
        source_distribution_rows(
            source_path,
            ws_values,
            ws_formulas,
            config,
            formula_evaluator,
            "Payroll Sales",
            [
                ("Trend House", 32, 7),
                ("OG Specialty USA", None, None),
                ("Online Lux", None, None),
                ("Online", 32, 9),
                ("DTC", 32, 8),
            ],
        )
    )
    rows.extend(
        source_distribution_rows(
            source_path,
            ws_values,
            ws_formulas,
            config,
            formula_evaluator,
            "Payroll Corp",
            [
                ("Corp", 61, 5),
                ("Art", 46, 5),
                ("IT", 37, 5),
                ("Production", 21, 5),
            ],
        )
    )

    payroll_corp_rows = [row for row in rows if row["block"] in {"Payroll Sales", "Payroll Corp"}]
    payroll_total = sum(row["amount"] or 0 for row in payroll_corp_rows)
    check_info = formula_cell_info(ws_values, ws_formulas, 68, 4, formula_evaluator, config)
    check_total = to_float(check_info["amount"])
    rows.append(
        derived_distribution_row(
            source_path,
            ws_values.title,
            "Payroll Corp",
            "Total",
            payroll_total,
            "derived_sum",
        )
    )
    rows.append(
        distribution_row_from_value_info(
            source_path,
            ws_values.title,
            "Payroll Corp",
            "Check Total",
            68,
            check_info,
            "D68",
        )
    )
    rows.append(
        derived_distribution_row(
            source_path,
            ws_values.title,
            "Payroll Corp",
            "Difference",
            payroll_total - check_total if check_total is not None else None,
            "derived_difference",
        )
    )

    lital_rows = source_distribution_rows(
        source_path,
        ws_values,
        ws_formulas,
        config,
        formula_evaluator,
        "Lital Allocation in G&A Exp",
        [
            ("DTC", 55, 13),
            ("ONLINE", 56, 13),
            ("TH", 57, 13),
            ("CORP", 59, 13),
        ],
    )
    rows.extend(lital_rows)
    rows.append(
        derived_distribution_row(
            source_path,
            ws_values.title,
            "Lital Allocation in G&A Exp",
            "Total",
            sum(row["amount"] or 0 for row in lital_rows),
            "derived_sum",
        )
    )
    return rows


def source_distribution_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: PayrollJournalConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    block: str,
    cells: list[tuple[str, int | None, int | None]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for label, row_idx, col_idx in cells:
        if row_idx is None or col_idx is None:
            rows.append(derived_distribution_row(source_path, ws_values.title, block, label, 0.0, "static_zero"))
            continue
        value_info = formula_cell_info(
            ws_values,
            ws_formulas,
            row_idx,
            col_idx,
            formula_evaluator,
            config,
        )
        rows.append(
            distribution_row_from_value_info(
                source_path,
                ws_values.title,
                block,
                label,
                row_idx,
                value_info,
                ws_values.cell(row_idx, col_idx).coordinate,
            )
        )
    return rows


def distribution_row_from_value_info(
    source_path: Path,
    sheet_name: str,
    block: str,
    label: str,
    row_idx: int,
    value_info: dict[str, Any],
    source_cell: str,
) -> dict[str, Any]:
    return {
        "source_file": str(source_path),
        "sheet": sheet_name,
        "block": block,
        "row": row_idx,
        "label": label,
        "amount": to_float(value_info["amount"]),
        "is_total_row": bool(re.search(r"\btotal\b", label, flags=re.IGNORECASE)),
        "is_check_row": bool(re.search(r"\bcheck\b", label, flags=re.IGNORECASE)),
        "is_difference_row": label.lower() == "difference",
        "formula": value_info["formula"],
        "cached_amount": to_float(value_info["cached_amount"]),
        "calculated_amount": to_float(value_info["calculated_amount"]),
        "calculation_status": value_info["calculation_status"],
        "calculation_detail": value_info["calculation_detail"],
        "source_cell": source_cell,
        "source_kind": "payroll_sheet",
    }


def derived_distribution_row(
    source_path: Path,
    sheet_name: str,
    block: str,
    label: str,
    amount: float | None,
    source_kind: str,
) -> dict[str, Any]:
    return {
        "source_file": str(source_path),
        "sheet": sheet_name,
        "block": block,
        "row": None,
        "label": label,
        "amount": amount,
        "is_total_row": bool(re.search(r"\btotal\b", label, flags=re.IGNORECASE)),
        "is_check_row": bool(re.search(r"\bcheck\b", label, flags=re.IGNORECASE)),
        "is_difference_row": label.lower() == "difference",
        "formula": None,
        "cached_amount": amount,
        "calculated_amount": None,
        "calculation_status": "derived",
        "calculation_detail": None,
        "source_cell": None,
        "source_kind": source_kind,
    }


def inferred_blank_distribution_label(previous_label: str | None, formula: str | None) -> str:
    if previous_label and "check" in previous_label.lower():
        return "Difference"
    if formula and "-" in formula:
        return "Difference"
    return "Total"


def formula_cell_info(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    col_idx: int,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    config: PayrollJournalConfig,
    use_calculated_formula_values: bool | None = None,
) -> dict[str, Any]:
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    calculated_value = None
    status = "not_formula"
    detail = None
    amount = cached_value

    if formula_evaluator is not None and formula_cell.data_type == "f":
        calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
        calculated_value = calculation.value
        status = calculation.status
        detail = calculation.detail
        should_use_calculated = (
            config.use_calculated_formula_values
            if use_calculated_formula_values is None
            else use_calculated_formula_values
        )
        if (
            should_use_calculated
            and calculation.status == "ok"
            and not is_formula_sentinel(calculation.value)
        ):
            amount = calculation.value

    return {
        "amount": amount,
        "formula": formula,
        "cached_amount": cached_value,
        "calculated_amount": calculated_value,
        "calculation_status": status,
        "calculation_detail": detail,
    }


def validate_extraction(
    employee_rows: list[dict[str, Any]],
    allocation_rows: list[dict[str, Any]],
    distribution_rows: list[dict[str, Any]],
) -> list[str]:
    warnings: list[str] = []
    if not employee_rows:
        warnings.append("No employee payroll rows parsed")
    if not allocation_rows:
        warnings.append("No employee allocation rows parsed")
    if not distribution_rows:
        warnings.append("No payroll distribution rows parsed")
    mismatches = [
        row
        for row in employee_rows
        if row["allocation_difference"] is not None and abs(row["allocation_difference"]) > 0.02
    ]
    if mismatches:
        warnings.append(f"{len(mismatches)} employee rows have allocation differences over 0.02")
    return warnings


def is_employee_row(ws: Worksheet, row_idx: int, config: PayrollJournalConfig) -> bool:
    return (
        is_number(ws.cell(row_idx, config.employee_code_column).value)
        and bool(normalize_text(ws.cell(row_idx, config.employee_name_column).value))
        and is_number(ws.cell(row_idx, config.gross_pay_column).value)
    )


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def is_number(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return False
        try:
            float(text)
        except ValueError:
            return False
        return True
    return False


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def matches_any(value: str | None, patterns: list[str]) -> bool:
    if not value:
        return False
    return any(re.search(pattern, value, flags=re.IGNORECASE) for pattern in patterns)


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(PayrollJournalConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
