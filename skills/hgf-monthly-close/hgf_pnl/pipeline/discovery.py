from __future__ import annotations

from datetime import datetime, timezone
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field


class PackageFile(BaseModel):
    path: str
    relative_path: str
    file_name: str
    suffix: str
    size_bytes: int
    modified_at: str
    role: str
    document_type: str
    extractor: str | None = None
    writer: str | None = None
    confidence: float
    reasons: list[str] = Field(default_factory=list)
    metadata: dict[str, Any] = Field(default_factory=dict)


class PackageDiscovery(BaseModel):
    root_path: str
    discovered_at: str
    files: list[PackageFile]
    warnings: list[str] = Field(default_factory=list)

    def by_extractor(self, extractor: str) -> list[PackageFile]:
        return [file for file in self.files if file.extractor == extractor]

    def by_role(self, role: str) -> list[PackageFile]:
        return [file for file in self.files if file.role == role]

    def to_json_file(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(self.model_dump_json(indent=2) + "\n", encoding="utf-8")


def discover_package(
    root_path: Path,
    *,
    include_zone_identifier: bool = False,
    include_temp_files: bool = False,
    inspect_workbooks: bool = False,
) -> PackageDiscovery:
    root_path = root_path.resolve()
    warnings: list[str] = []
    files: list[PackageFile] = []
    for path in sorted(root_path.rglob("*")):
        if not path.is_file():
            continue
        if should_skip_file(path, include_zone_identifier, include_temp_files):
            continue
        try:
            metadata = inspect_file(path, inspect_workbooks)
            files.append(classify_file(root_path, path, metadata))
        except Exception as exc:
            warnings.append(f"{path}: {type(exc).__name__}: {exc}")
            files.append(classify_file(root_path, path, {"inspection_error": str(exc)}))

    return PackageDiscovery(
        root_path=str(root_path),
        discovered_at=utc_now(),
        files=files,
        warnings=warnings,
    )


def should_skip_file(path: Path, include_zone_identifier: bool, include_temp_files: bool) -> bool:
    if not include_zone_identifier and path.name.endswith(":Zone.Identifier"):
        return True
    if not include_temp_files and path.name.startswith("~$"):
        return True
    if any(part.startswith(".") for part in path.parts):
        return True
    return False


def inspect_file(path: Path, inspect_workbooks: bool) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    if path.suffix.lower() in {".xlsx", ".xlsm"} and inspect_workbooks:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            metadata["sheet_names"] = list(workbook.sheetnames)
        finally:
            workbook.close()
    return metadata


def classify_file(root_path: Path, path: Path, metadata: dict[str, Any] | None = None) -> PackageFile:
    metadata = metadata or {}
    normalized = normalize_key(path.name)
    suffix = path.suffix.lower()
    reasons: list[str] = []
    role = "unknown"
    document_type = "unknown"
    extractor: str | None = None
    writer: str | None = None
    confidence = 0.1

    if suffix in {".xlsx", ".xlsm"}:
        document_type = "workbook"
    elif suffix == ".pdf":
        document_type = "pdf"
    elif suffix in {".docx", ".doc"}:
        document_type = "document"
    elif suffix in {".html", ".htm"}:
        document_type = "html"

    if all_words(normalized, ["profit", "loss", "dept"]):
        role, extractor, confidence = "source_input", "pl_by_dept", 0.95
        reasons.append("filename contains profit/loss/dept")
    elif "chargeback" in normalized and suffix == ".pdf":
        role, extractor, confidence = "source_input", "chargeback_pdf", 0.95
        reasons.append("chargeback PDF filename")
    elif all_words(normalized, ["payroll", "journal"]):
        role, extractor, confidence = "source_input", "payroll_journal", 0.95
        reasons.append("filename contains payroll journal")
    elif normalized == "br info" or all_words(normalized, ["br", "info"]):
        role, extractor, confidence = "source_input", "br_info", 0.95
        reasons.append("BR Info manual override workbook")
    elif all_words(normalized, ["monthly", "revenue"]) or all_words(normalized, ["dtc", "ws", "revenue"]):
        role, extractor, confidence = "source_input", "monthly_revenue", 0.95
        reasons.append("monthly revenue workbook filename")
    elif all_words(normalized, ["division", "cogs"]):
        role, extractor, confidence = "source_input", "division_cogs", 0.95
        reasons.append("division COGS workbook filename")
    elif "revenue report" in normalized and normalized.startswith("th "):
        role, extractor, confidence = "source_input", "th_revenue", 0.95
        reasons.append("Trend House revenue report filename")
    elif "addbacks" in normalized and suffix == ".pdf":
        role, confidence = "instruction", 0.9
        document_type = "instruction_pdf"
        reasons.append("addbacks instruction PDF")
    elif all_words(normalized, ["hgf", "gl"]) and "done" in normalized:
        role, extractor, confidence = "source_input", "addbacks_gl", 0.95
        reasons.append("reviewed GL workbook marked DONE")
    elif all_words(normalized, ["hgf", "gl"]):
        role, confidence = "supporting_input", 0.75
        reasons.append("original GL workbook")
    elif all_words(normalized, ["hgf", "consolidated"]) and "template" in normalized:
        role, writer, confidence = "template", "consolidated_pnl", 0.95
        document_type = "template_workbook"
        reasons.append("consolidated P&L template workbook")
    elif all_words(normalized, ["hgf", "consolidated"]):
        role, writer, confidence = "deliverable_or_prior_output", "consolidated_pnl", 0.85
        document_type = "deliverable_workbook"
        reasons.append("consolidated P&L workbook")
    elif "budget" in normalized:
        role, confidence = "supporting_input", 0.65
        reasons.append("budget support workbook")
    elif "blue" in normalized or "prior" in normalized:
        role, confidence = "supporting_input", 0.55
        reasons.append("prior-period support workbook")
    elif suffix in {".xlsx", ".xlsm", ".pdf", ".docx", ".doc", ".html", ".htm"}:
        role, confidence = "supporting_input", 0.35
        reasons.append("recognized document type but no specific extractor match")
    else:
        reasons.append("no classification rule matched")

    sheet_names = metadata.get("sheet_names") or []
    if sheet_names:
        role, extractor, writer, confidence, reasons = refine_from_sheet_names(
            sheet_names,
            role,
            extractor,
            writer,
            confidence,
            reasons,
        )

    stat = path.stat()
    return PackageFile(
        path=str(path.resolve()),
        relative_path=str(path.resolve().relative_to(root_path)),
        file_name=path.name,
        suffix=suffix,
        size_bytes=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        role=role,
        document_type=document_type,
        extractor=extractor,
        writer=writer,
        confidence=round(confidence, 3),
        reasons=reasons,
        metadata=metadata,
    )


def refine_from_sheet_names(
    sheet_names: list[str],
    role: str,
    extractor: str | None,
    writer: str | None,
    confidence: float,
    reasons: list[str],
) -> tuple[str, str | None, str | None, float, list[str]]:
    normalized_sheets = [normalize_key(sheet) for sheet in sheet_names]
    sheet_text = " ".join(normalized_sheets)
    if "raw data master file" in sheet_text and "march 2026 full" in sheet_text:
        if writer is None:
            writer = "consolidated_pnl"
        if role == "unknown":
            role = "deliverable_or_prior_output"
        confidence = max(confidence, 0.85)
        reasons.append("workbook has consolidated P&L raw/full tabs")
    if "new month" in normalized_sheets:
        if extractor is None:
            role, extractor = "source_input", "addbacks_gl"
        confidence = max(confidence, 0.85)
        reasons.append("workbook has reviewed GL NEW MONTH sheet")
    if "payroll" in normalized_sheets and "payroll distribution" in normalized_sheets:
        if extractor is None:
            role, extractor = "source_input", "payroll_journal"
        confidence = max(confidence, 0.85)
        reasons.append("workbook has payroll and payroll distribution sheets")
    return role, extractor, writer, confidence, reasons


def write_discovery(path: Path, discovery: PackageDiscovery) -> None:
    discovery.to_json_file(path)


def load_discovery(path: Path) -> PackageDiscovery:
    return PackageDiscovery.model_validate_json(path.read_text(encoding="utf-8"))


def all_words(text: str, words: list[str]) -> bool:
    return all(word in text for word in words)


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def discovery_summary(discovery: PackageDiscovery) -> dict[str, Any]:
    by_extractor: dict[str, int] = {}
    by_role: dict[str, int] = {}
    for file in discovery.files:
        by_role[file.role] = by_role.get(file.role, 0) + 1
        if file.extractor:
            by_extractor[file.extractor] = by_extractor.get(file.extractor, 0) + 1
        if file.writer:
            by_extractor[f"writer:{file.writer}"] = by_extractor.get(f"writer:{file.writer}", 0) + 1
    return {
        "root_path": discovery.root_path,
        "file_count": len(discovery.files),
        "by_role": dict(sorted(by_role.items())),
        "by_extractor": dict(sorted(by_extractor.items())),
        "warnings": discovery.warnings,
    }


def discovery_summary_json(discovery: PackageDiscovery) -> str:
    return json.dumps(discovery_summary(discovery), indent=2) + "\n"
