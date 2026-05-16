from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import shutil
from typing import Any

from openpyxl import load_workbook


MASTER_SHEET = "RAW DATA_Master File"


@dataclass(frozen=True)
class TemplatePatch:
    sheet_name: str
    cell: str
    old_value: Any
    new_value: Any
    note: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "cell": self.cell,
            "old_value": self.old_value,
            "new_value": self.new_value,
            "note": self.note,
        }


@dataclass(frozen=True)
class TemplatePatchResult:
    source_path: Path
    output_path: Path
    full_sheet_name: str
    patches: list[TemplatePatch]

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_path": str(self.source_path),
            "output_path": str(self.output_path),
            "full_sheet_name": self.full_sheet_name,
            "patches": [patch.to_dict() for patch in self.patches],
        }


def patch_consolidated_template(
    source_path: Path,
    output_path: Path,
    *,
    full_sheet_name: str | None = None,
    unhide_all_sheets: bool = False,
) -> TemplatePatchResult:
    source_path = source_path.expanduser().resolve()
    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path != source_path:
        shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    try:
        full_sheet_name = full_sheet_name or detect_full_sheet(workbook.sheetnames)
        if full_sheet_name not in workbook.sheetnames:
            raise ValueError(f"Full report sheet not found: {full_sheet_name!r}")
        if MASTER_SHEET not in workbook.sheetnames:
            raise ValueError(f"Raw master sheet not found: {MASTER_SHEET!r}")

        full = workbook[full_sheet_name]
        master = workbook[MASTER_SHEET]
        patches = [
            apply_patch(full, "AA31", 0, "Break Online LUX travel duplicate of OG-DTC travel."),
            apply_patch(
                full,
                "AW52",
                "=+'RAW DATA_Master File'!B23",
                "Route Merchant Account Fees through raw-data value instead of hardcoded literal.",
            ),
            apply_patch(full, "BH52", 0, "Clear APA phantom merchant-fee literal."),
            apply_patch(full, "AL55", "=0.25*EB55", "Remove phantom Art Assets offset."),
            apply_patch(full, "BH55", "=0.2*EB55", "Make Art Assets channel allocation sum to 100%."),
            apply_patch(
                master,
                "B100",
                "=+'RAW DATA_COGS & Freight'!G5+'RAW DATA_COGS & Freight'!K5",
                "Include Online-USA shipping in Online channel shipping total.",
            ),
        ]

        if unhide_all_sheets:
            for sheet in workbook.worksheets:
                sheet.sheet_state = "visible"

        workbook.save(output_path)
    finally:
        workbook.close()

    return TemplatePatchResult(
        source_path=source_path,
        output_path=output_path,
        full_sheet_name=full_sheet_name,
        patches=patches,
    )


def detect_full_sheet(sheet_names: list[str]) -> str:
    candidates = [name for name in sheet_names if "full" in name.lower()]
    if not candidates:
        raise ValueError("Could not find a sheet whose name contains 'FULL'")
    return candidates[0]


def apply_patch(ws: Any, cell: str, new_value: Any, note: str) -> TemplatePatch:
    old_value = ws[cell].value
    ws[cell] = new_value
    return TemplatePatch(
        sheet_name=ws.title,
        cell=cell,
        old_value=old_value,
        new_value=new_value,
        note=note,
    )
