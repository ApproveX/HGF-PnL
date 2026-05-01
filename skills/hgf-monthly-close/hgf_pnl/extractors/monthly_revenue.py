from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any, Literal

import polars as pl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from rapidfuzz import fuzz

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


SheetRole = Literal["summary", "shopify", "refunds", "coupons"]


ROLE_KEYWORDS: dict[SheetRole, list[str]] = {
    "summary": ["revenue", "summary"],
    "shopify": ["shopify"],
    "refunds": ["refund"],
    "coupons": ["coupon"],
}


ROLE_HEADER_ALIASES: dict[SheetRole, dict[str, list[str]]] = {
    "shopify": {
        "day": ["day", "date"],
        "order_name": ["order name", "order"],
        "customer_name": ["customer name", "customer"],
        "customer_email": ["customer email", "email"],
        "gross_sales": ["gross sales"],
        "orders": ["orders"],
        "quantity_ordered_per_order": ["quantity ordered per order", "quantity ordered"],
        "average_order_value": ["average order value"],
        "quantity_returned": ["quantity returned"],
        "net_sales": ["net sales"],
        "channel": ["dtc or ws", "channel"],
        "method": ["method"],
    },
    "refunds": {
        "date": ["date"],
        "year": ["year"],
        "month": ["month"],
        "requested_by": ["who is requesting", "requesting"],
        "order_number": ["order #", "order number", "order"],
        "division": ["og", "division"],
        "amount": ["amount"],
        "refund_category": ["refund category"],
        "return_reason": ["return reason"],
        "return_sku": ["return sku", "sku"],
        "model_number": ["model #", "model number"],
        "image": ["image"],
        "size": ["size"],
        "acrylic": ["acrylic"],
        "embellishment": ["embellishment"],
        "notes": ["notes feedback", "notes / feedback", "notes"],
        "saved": ["saved"],
        "payment_method": ["payment method"],
        "pp_customer_email": ["pp customer email", "customer email"],
        "jason_approval": ["jason approval"],
        "jason_comments": ["jason comments"],
        "notified_date": ["refund date customer notified date", "customer notified date"],
    },
    "coupons": {
        "order": ["order"],
        "date": ["date"],
        "customer": ["customer"],
        "payment_status": ["payment status"],
        "fulfillment_status": ["fulfillment status"],
        "items": ["items"],
        "total": ["total"],
        "channel": ["channel"],
        "delivery_status": ["delivery status"],
        "delivery_method": ["delivery method"],
    },
    "summary": {},
}


ROLE_REQUIRED_COLUMNS: dict[SheetRole, set[str]] = {
    "summary": set(),
    "shopify": {"day", "order_name", "net_sales", "channel"},
    "refunds": {"date", "order_number", "division", "amount", "refund_category"},
    "coupons": {"order", "date", "total"},
}


class MonthlyRevenueConfig(BaseModel):
    """Agent-adjustable extraction rules for DTC/WS monthly revenue workbooks."""

    sheet_roles: dict[SheetRole, list[str]] = Field(default_factory=lambda: ROLE_KEYWORDS.copy())
    column_aliases: dict[SheetRole, dict[str, list[str]]] = Field(default_factory=lambda: ROLE_HEADER_ALIASES.copy())
    max_header_scan_rows: int = 8
    fuzzy_header_threshold: int = 86
    stop_after_blank_rows: int = 5
    include_refund_rows_without_amount: bool = True
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "MonthlyRevenueConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class MonthlyRevenueExtraction:
    path: Path
    sheets: dict[SheetRole, str]
    summary_rows: list[dict[str, Any]]
    sales_rows: list[dict[str, Any]]
    refund_rows: list[dict[str, Any]]
    coupon_rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    @property
    def summary(self) -> pl.DataFrame:
        return pl.DataFrame(self.summary_rows)

    @property
    def sales(self) -> pl.DataFrame:
        return pl.DataFrame(self.sales_rows)

    @property
    def refunds(self) -> pl.DataFrame:
        return pl.DataFrame(self.refund_rows)

    @property
    def coupons(self) -> pl.DataFrame:
        return pl.DataFrame(self.coupon_rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheets": self.sheets,
            "warnings": self.warnings,
            "summary_rows": self.summary_rows,
            "sales_rows": self.sales_rows,
            "refund_rows": self.refund_rows,
            "coupon_rows": self.coupon_rows,
        }


def extract_monthly_revenue(
    path: Path,
    config: MonthlyRevenueConfig | None = None,
) -> MonthlyRevenueExtraction:
    config = config or MonthlyRevenueConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        role_to_sheet = choose_role_sheets(workbook_values.sheetnames, config)
        warnings: list[str] = []
        missing_roles = set(ROLE_KEYWORDS) - set(role_to_sheet)
        if missing_roles:
            warnings.append(f"Missing expected sheets for roles: {sorted(missing_roles)}")

        summary_rows: list[dict[str, Any]] = []
        sales_rows: list[dict[str, Any]] = []
        refund_rows: list[dict[str, Any]] = []
        coupon_rows: list[dict[str, Any]] = []

        if "summary" in role_to_sheet:
            sheet_name = role_to_sheet["summary"]
            summary_rows = extract_summary_rows(path, workbook_values[sheet_name], workbook_formulas[sheet_name], config, formula_evaluator)

        if "shopify" in role_to_sheet:
            sheet_name = role_to_sheet["shopify"]
            sales_rows = extract_shopify_rows(path, workbook_values[sheet_name], workbook_formulas[sheet_name], config, formula_evaluator)

        if "refunds" in role_to_sheet:
            sheet_name = role_to_sheet["refunds"]
            refund_rows = extract_refund_rows(path, workbook_values[sheet_name], workbook_formulas[sheet_name], config, formula_evaluator)

        if "coupons" in role_to_sheet:
            sheet_name = role_to_sheet["coupons"]
            coupon_rows = extract_coupon_rows(path, workbook_values[sheet_name], workbook_formulas[sheet_name], config, formula_evaluator)

        warnings.extend(validate_extraction(summary_rows, sales_rows, refund_rows, coupon_rows))
        return MonthlyRevenueExtraction(
            path=path,
            sheets=role_to_sheet,
            summary_rows=summary_rows,
            sales_rows=sales_rows,
            refund_rows=refund_rows,
            coupon_rows=coupon_rows,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def choose_role_sheets(sheet_names: list[str], config: MonthlyRevenueConfig) -> dict[SheetRole, str]:
    chosen: dict[SheetRole, str] = {}
    used: set[str] = set()
    for role, keywords in config.sheet_roles.items():
        scored: list[tuple[int, str]] = []
        lowered_keywords = [normalize_key(keyword) for keyword in keywords]
        for sheet_name in sheet_names:
            if sheet_name in used:
                continue
            normalized = normalize_key(sheet_name)
            exact_score = sum(keyword in normalized for keyword in lowered_keywords) * 100
            fuzzy_score = max(
                (fuzz.partial_ratio(keyword, normalized) for keyword in lowered_keywords),
                default=0,
            )
            scored.append((exact_score + int(fuzzy_score), sheet_name))
        scored.sort(reverse=True)
        if scored and scored[0][0] >= config.fuzzy_header_threshold:
            chosen[role] = scored[0][1]
            used.add(scored[0][1])
    return chosen


def extract_summary_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: MonthlyRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_section: str | None = None
    current_metric: str | None = None

    for row_idx in range(1, ws_values.max_row + 1):
        label = normalize_text(ws_values.cell(row_idx, 1).value)
        amount_cell_value = ws_values.cell(row_idx, 2).value
        amount_text = normalize_text(amount_cell_value)
        if not label and not amount_text:
            continue
        if label and not amount_text:
            current_section = label
            current_metric = None
            continue
        if normalize_key(label) == "row labels":
            current_metric = label_or_default(amount_cell_value, "Amount")
            continue
        if current_section is None:
            continue

        value_info = formula_cell_info(ws_values, ws_formulas, row_idx, 2, formula_evaluator, config)
        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "section": current_section,
                "metric": current_metric,
                "row": row_idx,
                "label": label,
                "amount": to_float(value_info["amount"]),
                "is_total_row": bool(re.search(r"\btotal\b", label, flags=re.IGNORECASE)),
                "source_cell": f"B{row_idx}",
                "formula": value_info["formula"],
                "cached_amount": to_float(value_info["cached_value"]),
                "calculated_amount": to_float(value_info["calculated_value"]),
                "calculation_status": value_info["calculation_status"],
                "calculation_detail": value_info["calculation_detail"],
            }
        )
    return rows


def extract_shopify_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: MonthlyRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    header_row, column_map, source_headers = detect_header(ws_values, "shopify", config)
    rows: list[dict[str, Any]] = []
    for row_idx in iter_data_rows(ws_values, header_row, column_map, config):
        formula_status: dict[str, str] = {}
        formula_detail: dict[str, str | None] = {}
        gross_sales = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "gross_sales", config, formula_evaluator, formula_status, formula_detail)
        orders = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "orders", config, formula_evaluator, formula_status, formula_detail)
        quantity_ordered = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "quantity_ordered_per_order", config, formula_evaluator, formula_status, formula_detail)
        average_order_value = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "average_order_value", config, formula_evaluator, formula_status, formula_detail)
        quantity_returned = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "quantity_returned", config, formula_evaluator, formula_status, formula_detail)
        net_sales = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "net_sales", config, formula_evaluator, formula_status, formula_detail)

        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "row": row_idx,
                "day": normalize_date_value(cell_value(ws_values, row_idx, column_map, "day")),
                "order_name": normalize_text(cell_value(ws_values, row_idx, column_map, "order_name")),
                "customer_name": normalize_text(cell_value(ws_values, row_idx, column_map, "customer_name")),
                "customer_email": normalize_text(cell_value(ws_values, row_idx, column_map, "customer_email")),
                "gross_sales": to_float(gross_sales),
                "orders": to_float(orders),
                "quantity_ordered_per_order": to_float(quantity_ordered),
                "average_order_value": to_float(average_order_value),
                "quantity_returned": to_float(quantity_returned),
                "net_sales": to_float(net_sales),
                "channel": normalize_text(cell_value(ws_values, row_idx, column_map, "channel")),
                "method": normalize_text(cell_value(ws_values, row_idx, column_map, "method")),
                "source_headers": json.dumps(source_headers, sort_keys=True),
                "formula_status": json.dumps(formula_status, sort_keys=True),
                "formula_detail": json.dumps(formula_detail, sort_keys=True),
            }
        )
    return rows


def extract_refund_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: MonthlyRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    header_row, column_map, source_headers = detect_header(ws_values, "refunds", config)
    rows: list[dict[str, Any]] = []
    for row_idx in iter_data_rows(ws_values, header_row, column_map, config):
        formula_status: dict[str, str] = {}
        formula_detail: dict[str, str | None] = {}
        amount = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "amount", config, formula_evaluator, formula_status, formula_detail)
        amount_value = to_float(amount)
        if amount_value is None and not config.include_refund_rows_without_amount:
            continue

        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "row": row_idx,
                "date": normalize_date_value(cell_value(ws_values, row_idx, column_map, "date")),
                "year": to_int(cell_value(ws_values, row_idx, column_map, "year")),
                "month": to_int(cell_value(ws_values, row_idx, column_map, "month")),
                "requested_by": normalize_text(cell_value(ws_values, row_idx, column_map, "requested_by")),
                "order_number": normalize_order_number(cell_value(ws_values, row_idx, column_map, "order_number")),
                "division": normalize_text(cell_value(ws_values, row_idx, column_map, "division")),
                "amount": amount_value,
                "has_amount": amount_value is not None,
                "refund_category": normalize_text(cell_value(ws_values, row_idx, column_map, "refund_category")),
                "return_reason": normalize_text(cell_value(ws_values, row_idx, column_map, "return_reason")),
                "return_sku": normalize_text(cell_value(ws_values, row_idx, column_map, "return_sku")),
                "model_number": normalize_text(cell_value(ws_values, row_idx, column_map, "model_number")),
                "image": normalize_text(cell_value(ws_values, row_idx, column_map, "image")),
                "size": normalize_text(cell_value(ws_values, row_idx, column_map, "size")),
                "acrylic": normalize_text(cell_value(ws_values, row_idx, column_map, "acrylic")),
                "embellishment": normalize_text(cell_value(ws_values, row_idx, column_map, "embellishment")),
                "notes": normalize_text(cell_value(ws_values, row_idx, column_map, "notes")),
                "saved": normalize_text(cell_value(ws_values, row_idx, column_map, "saved")),
                "payment_method": normalize_text(cell_value(ws_values, row_idx, column_map, "payment_method")),
                "pp_customer_email": normalize_text(cell_value(ws_values, row_idx, column_map, "pp_customer_email")),
                "jason_approval": normalize_text(cell_value(ws_values, row_idx, column_map, "jason_approval")),
                "jason_comments": normalize_text(cell_value(ws_values, row_idx, column_map, "jason_comments")),
                "notified_date": normalize_text(cell_value(ws_values, row_idx, column_map, "notified_date")),
                "source_headers": json.dumps(source_headers, sort_keys=True),
                "formula_status": json.dumps(formula_status, sort_keys=True),
                "formula_detail": json.dumps(formula_detail, sort_keys=True),
            }
        )
    return rows


def extract_coupon_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: MonthlyRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    header_row, column_map, source_headers = detect_header(ws_values, "coupons", config)
    rows: list[dict[str, Any]] = []
    for row_idx in iter_data_rows(ws_values, header_row, column_map, config):
        formula_status: dict[str, str] = {}
        formula_detail: dict[str, str | None] = {}
        total = effective_cell_value(ws_values, ws_formulas, row_idx, column_map, "total", config, formula_evaluator, formula_status, formula_detail)

        rows.append(
            {
                "source_file": str(source_path),
                "sheet": ws_values.title,
                "row": row_idx,
                "order": normalize_text(cell_value(ws_values, row_idx, column_map, "order")),
                "date": normalize_date_value(cell_value(ws_values, row_idx, column_map, "date")),
                "customer": normalize_text(cell_value(ws_values, row_idx, column_map, "customer")),
                "payment_status": normalize_text(cell_value(ws_values, row_idx, column_map, "payment_status")),
                "fulfillment_status": normalize_text(cell_value(ws_values, row_idx, column_map, "fulfillment_status")),
                "items": normalize_text(cell_value(ws_values, row_idx, column_map, "items")),
                "total": to_float(total),
                "channel": normalize_text(cell_value(ws_values, row_idx, column_map, "channel")),
                "delivery_status": normalize_text(cell_value(ws_values, row_idx, column_map, "delivery_status")),
                "delivery_method": normalize_text(cell_value(ws_values, row_idx, column_map, "delivery_method")),
                "source_headers": json.dumps(source_headers, sort_keys=True),
                "formula_status": json.dumps(formula_status, sort_keys=True),
                "formula_detail": json.dumps(formula_detail, sort_keys=True),
            }
        )
    return rows


def detect_header(
    ws: Worksheet,
    role: SheetRole,
    config: MonthlyRevenueConfig,
) -> tuple[int, dict[str, int], dict[str, str]]:
    candidates: list[tuple[int, int, dict[str, int], dict[str, str]]] = []
    for row_idx in range(1, min(ws.max_row, config.max_header_scan_rows) + 1):
        headers = {
            col_idx: normalize_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
            if normalize_text(ws.cell(row_idx, col_idx).value)
        }
        column_map, source_headers = map_headers(headers, role, config)
        required_hits = len(ROLE_REQUIRED_COLUMNS[role] & set(column_map))
        score = required_hits * 20 + len(column_map) * 5 + min(len(headers), 20)
        if column_map:
            candidates.append((score, row_idx, column_map, source_headers))
    if not candidates:
        raise ValueError(f"Could not detect header row for {ws.title!r}")
    candidates.sort(reverse=True)
    _, row_idx, column_map, source_headers = candidates[0]
    missing = ROLE_REQUIRED_COLUMNS[role] - set(column_map)
    if missing:
        raise ValueError(f"Missing expected columns in {ws.title!r}: {sorted(missing)}")
    return row_idx, column_map, source_headers


def map_headers(
    headers: dict[int, str],
    role: SheetRole,
    config: MonthlyRevenueConfig,
) -> tuple[dict[str, int], dict[str, str]]:
    mapped: dict[str, int] = {}
    source_headers: dict[str, str] = {}
    aliases = config.column_aliases[role]
    for col_idx, header in headers.items():
        canonical = match_header(header, aliases, config.fuzzy_header_threshold)
        if canonical and canonical not in mapped:
            mapped[canonical] = col_idx
            source_headers[canonical] = header
    return mapped, source_headers


def match_header(header: str, aliases: dict[str, list[str]], threshold: int) -> str | None:
    normalized_header = normalize_key(header)
    best: tuple[int, str] | None = None
    for canonical, candidates in aliases.items():
        for candidate in candidates:
            normalized_candidate = normalize_key(candidate)
            if normalized_header == normalized_candidate:
                return canonical
            if normalized_candidate in normalized_header or normalized_header in normalized_candidate:
                score = 95
            else:
                score = int(fuzz.token_sort_ratio(normalized_header, normalized_candidate))
            if best is None or score > best[0]:
                best = (score, canonical)
    if best and best[0] >= threshold:
        return best[1]
    return None


def iter_data_rows(
    ws: Worksheet,
    header_row: int,
    column_map: dict[str, int],
    config: MonthlyRevenueConfig,
) -> list[int]:
    rows: list[int] = []
    blank_run = 0
    columns = sorted(set(column_map.values()))
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in columns]
        if all(value in (None, "") for value in values):
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            continue
        blank_run = 0
        rows.append(row_idx)
    return rows


def effective_cell_value(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    column_map: dict[str, int],
    field: str,
    config: MonthlyRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    formula_status: dict[str, str],
    formula_detail: dict[str, str | None],
) -> Any:
    col_idx = column_map.get(field)
    if col_idx is None:
        return None
    value_info = formula_cell_info(ws_values, ws_formulas, row_idx, col_idx, formula_evaluator, config)
    if value_info["formula"] is not None:
        formula_status[field] = value_info["calculation_status"]
        formula_detail[field] = value_info["calculation_detail"]
    return value_info["amount"]


def formula_cell_info(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    col_idx: int,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    config: MonthlyRevenueConfig,
) -> dict[str, Any]:
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    calculated_value = None
    status = "not_formula"
    detail = None
    amount = cached_value

    if formula_evaluator is not None and formula_cell.data_type == "f":
        calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
        calculated_value = calculation.value
        status = calculation.status
        detail = calculation.detail
        if (
            config.use_calculated_formula_values
            and calculation.status == "ok"
            and not is_formula_sentinel(calculation.value)
        ):
            amount = calculation.value

    return {
        "amount": amount,
        "formula": formula,
        "cached_value": cached_value,
        "calculated_value": calculated_value,
        "calculation_status": status,
        "calculation_detail": detail,
    }


def cell_value(ws: Worksheet, row_idx: int, column_map: dict[str, int], field: str) -> Any:
    col_idx = column_map.get(field)
    if col_idx is None:
        return None
    return ws.cell(row_idx, col_idx).value


def validate_extraction(
    summary_rows: list[dict[str, Any]],
    sales_rows: list[dict[str, Any]],
    refund_rows: list[dict[str, Any]],
    coupon_rows: list[dict[str, Any]],
) -> list[str]:
    warnings: list[str] = []
    if not summary_rows:
        warnings.append("No summary rows parsed")
    if not sales_rows:
        warnings.append("No Shopify sales rows parsed")
    if not refund_rows:
        warnings.append("No refund rows parsed")
    if not coupon_rows:
        warnings.append("No coupon rows parsed")
    return warnings


def label_or_default(value: Any, default: str) -> str:
    normalized = normalize_text(value)
    return normalized or default


def normalize_date_value(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime | date):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return normalize_text(value)


def normalize_order_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_text(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def to_int(value: Any) -> int | None:
    number = to_float(value)
    if number is None:
        return None
    return int(number)


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(MonthlyRevenueConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
