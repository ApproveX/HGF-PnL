from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


class DivisionCOGSConfig(BaseModel):
    """Agent-adjustable extraction rules for Division COGS workbooks."""

    year_sheet_pattern: str = r"^20\d{2}$"
    partner_detail_sheet_pattern: str = r"^20\d{2}\s+Partner Details$"
    include_total_columns: bool = True
    include_zero_amounts: bool = False
    include_error_values: bool = True
    stop_after_blank_rows: int = 20
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "DivisionCOGSConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class DivisionCOGSExtraction:
    path: Path
    year_sheets: list[str]
    partner_detail_sheets: list[str]
    matrix_rows: list[dict[str, Any]]
    partner_detail_rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    @property
    def matrix(self) -> pl.DataFrame:
        return pl.DataFrame(self.matrix_rows, infer_schema_length=None)

    @property
    def partner_details(self) -> pl.DataFrame:
        return pl.DataFrame(self.partner_detail_rows, infer_schema_length=None)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "year_sheets": self.year_sheets,
            "partner_detail_sheets": self.partner_detail_sheets,
            "warnings": self.warnings,
            "matrix_rows": self.matrix_rows,
            "partner_detail_rows": self.partner_detail_rows,
        }


def extract_division_cogs(
    path: Path,
    config: DivisionCOGSConfig | None = None,
) -> DivisionCOGSExtraction:
    config = config or DivisionCOGSConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        year_sheets = [
            sheet
            for sheet in workbook_values.sheetnames
            if re.search(config.year_sheet_pattern, sheet, flags=re.IGNORECASE)
        ]
        partner_detail_sheets = [
            sheet
            for sheet in workbook_values.sheetnames
            if re.search(config.partner_detail_sheet_pattern, sheet, flags=re.IGNORECASE)
        ]

        matrix_rows: list[dict[str, Any]] = []
        partner_detail_rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        for sheet_name in year_sheets:
            try:
                matrix_rows.extend(
                    extract_year_matrix_sheet(
                        path,
                        workbook_values[sheet_name],
                        workbook_formulas[sheet_name],
                        config,
                        formula_evaluator,
                    )
                )
            except Exception as exc:
                warnings.append(f"{sheet_name}: {type(exc).__name__}: {exc}")

        for sheet_name in partner_detail_sheets:
            try:
                partner_detail_rows.extend(
                    extract_partner_detail_sheet(
                        path,
                        workbook_values[sheet_name],
                        workbook_formulas[sheet_name],
                        config,
                        formula_evaluator,
                    )
                )
            except Exception as exc:
                warnings.append(f"{sheet_name}: {type(exc).__name__}: {exc}")

        warnings.extend(validate_extraction(matrix_rows, partner_detail_rows, year_sheets, partner_detail_sheets))
        return DivisionCOGSExtraction(
            path=path,
            year_sheets=year_sheets,
            partner_detail_sheets=partner_detail_sheets,
            matrix_rows=matrix_rows,
            partner_detail_rows=partner_detail_rows,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def extract_year_matrix_sheet(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: DivisionCOGSConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    header_row, old_style = detect_year_header_row(ws_values)
    month_col = 1
    type_col = 2 if not old_style else None
    value_start_col = 2 if old_style else 3
    channels = contiguous_headers(ws_values, header_row, value_start_col)
    rows: list[dict[str, Any]] = []
    current_month: datetime | date | None = None
    blank_run = 0

    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        month_value = ws_values.cell(row_idx, month_col).value
        row_type = "COGS" if old_style else normalize_text(ws_values.cell(row_idx, type_col or 2).value)
        if isinstance(month_value, datetime | date):
            current_month = month_value
        elif normalize_text(month_value).lower().startswith("total"):
            current_month = None
            continue

        has_data = bool(row_type) or any(
            ws_values.cell(row_idx, col_idx).value not in (None, "")
            for col_idx in channels
        )
        if not has_data:
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            continue
        blank_run = 0

        if current_month is None or not row_type:
            continue

        for col_idx, channel in channels.items():
            if not config.include_total_columns and normalize_key(channel) == "total":
                continue
            value_info = formula_cell_info(ws_values, ws_formulas, row_idx, col_idx, formula_evaluator, config)
            amount = to_float(value_info["amount"])
            raw_value = normalize_text(value_info["amount"])
            if amount is None:
                if not config.include_error_values or raw_value in {"", "N/A"}:
                    continue
            elif amount == 0 and not config.include_zero_amounts:
                continue

            rows.append(
                {
                    "source_file": str(source_path),
                    "sheet": ws_values.title,
                    "year": year_from_month(current_month),
                    "month_num": month_from_value(current_month),
                    "month_name": month_name(month_from_value(current_month)),
                    "month": normalize_date_value(current_month),
                    "row": row_idx,
                    "type": row_type,
                    "channel": channel,
                    "amount": amount,
                    "raw_value": raw_value,
                    "is_total_column": normalize_key(channel) == "total",
                    "source_cell": f"{get_column_letter(col_idx)}{row_idx}",
                    "formula": value_info["formula"],
                    "cached_amount": to_float(value_info["cached_value"]),
                    "calculated_amount": to_float(value_info["calculated_value"]),
                    "calculation_status": value_info["calculation_status"],
                    "calculation_detail": value_info["calculation_detail"],
                }
            )
    return rows


def detect_year_header_row(ws: Worksheet) -> tuple[int, bool]:
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        first = normalize_key(ws.cell(row_idx, 1).value)
        second = normalize_key(ws.cell(row_idx, 2).value)
        if first == "month" and second == "type":
            return row_idx, False
        if first == "cogs" and second:
            return row_idx, True
    raise ValueError(f"Could not detect COGS year matrix header row in {ws.title!r}")


def contiguous_headers(ws: Worksheet, header_row: int, start_col: int) -> dict[int, str]:
    headers: dict[int, str] = {}
    for col_idx in range(start_col, ws.max_column + 1):
        header = normalize_text(ws.cell(header_row, col_idx).value)
        if not header:
            if headers:
                break
            continue
        if normalize_key(header) == "n a":
            continue
        headers[col_idx] = header
    return headers


def extract_partner_detail_sheet(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    config: DivisionCOGSConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    month_row, measure_row, month_by_col = detect_partner_header_rows(ws_values)
    rows: list[dict[str, Any]] = []
    blank_run = 0
    for row_idx in range(measure_row + 1, ws_values.max_row + 1):
        partner = normalize_text(ws_values.cell(row_idx, 1).value)
        if not partner:
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            continue
        blank_run = 0

        for col_idx, month_value in month_by_col.items():
            measure = normalize_partner_measure(ws_values.cell(measure_row, col_idx).value)
            if not measure:
                continue
            value_info = formula_cell_info(ws_values, ws_formulas, row_idx, col_idx, formula_evaluator, config)
            amount = to_float(value_info["amount"])
            raw_value = normalize_text(value_info["amount"])
            if amount is None:
                if not config.include_error_values or raw_value in {"", "N/A"}:
                    continue
            elif amount == 0 and not config.include_zero_amounts:
                continue
            rows.append(
                {
                    "source_file": str(source_path),
                    "sheet": ws_values.title,
                    "year": year_from_month(month_value),
                    "month_num": month_from_value(month_value),
                    "month_name": month_name(month_from_value(month_value)),
                    "month": normalize_date_value(month_value),
                    "row": row_idx,
                    "partner": partner,
                    "measure": measure,
                    "amount": amount,
                    "raw_value": raw_value,
                    "source_cell": f"{get_column_letter(col_idx)}{row_idx}",
                    "month_header_row": month_row,
                    "measure_header_row": measure_row,
                    "formula": value_info["formula"],
                    "cached_amount": to_float(value_info["cached_value"]),
                    "calculated_amount": to_float(value_info["calculated_value"]),
                    "calculation_status": value_info["calculation_status"],
                    "calculation_detail": value_info["calculation_detail"],
                }
            )
    return rows


def detect_partner_header_rows(ws: Worksheet) -> tuple[int, int, dict[int, datetime | date]]:
    best: tuple[int, int, dict[int, datetime | date]] | None = None
    for row_idx in range(1, min(ws.max_row, 6) + 1):
        month_starts = {
            col_idx: value
            for col_idx in range(1, ws.max_column + 1)
            if (value := parse_month_value(ws.cell(row_idx, col_idx).value)) is not None
        }
        if len(month_starts) < 1:
            continue
        if best is None or len(month_starts) > len(best[2]):
            best = (row_idx, row_idx + 1, month_starts)
    if best is None:
        raise ValueError(f"Could not detect partner detail month header row in {ws.title!r}")

    month_row, measure_row, month_starts = best
    sorted_starts = sorted(month_starts.items())
    month_by_col: dict[int, datetime | date] = {}
    for idx, (start_col, month_value) in enumerate(sorted_starts):
        next_start = sorted_starts[idx + 1][0] if idx + 1 < len(sorted_starts) else ws.max_column + 1
        for col_idx in range(start_col, next_start):
            if normalize_partner_measure(ws.cell(measure_row, col_idx).value):
                month_by_col[col_idx] = month_value
    return month_row, measure_row, month_by_col


def parse_month_value(value: Any) -> datetime | date | None:
    if isinstance(value, datetime | date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    match = re.fullmatch(r"(\d{1,2})/(\d{4})", text)
    if match:
        return date(int(match.group(2)), int(match.group(1)), 1)
    return None


def normalize_partner_measure(value: Any) -> str | None:
    normalized = normalize_key(value)
    if not normalized:
        return None
    if normalized in {"sum of cogs", "cogs"}:
        return "cogs"
    if normalized in {"sum of material cost", "sum of material", "material cost", "material"}:
        return "material_cost"
    if normalized in {"sum of labor cost", "sum of labor", "labor cost", "labor"}:
        return "labor_cost"
    return None


def formula_cell_info(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    col_idx: int,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    config: DivisionCOGSConfig,
) -> dict[str, Any]:
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    calculated_value = None
    status = "not_formula"
    detail = None
    amount = cached_value

    if formula_evaluator is not None and formula_cell.data_type == "f":
        calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
        calculated_value = calculation.value
        status = calculation.status
        detail = calculation.detail
        if (
            config.use_calculated_formula_values
            and calculation.status == "ok"
            and not is_formula_sentinel(calculation.value)
        ):
            amount = calculation.value

    return {
        "amount": amount,
        "formula": formula,
        "cached_value": cached_value,
        "calculated_value": calculated_value,
        "calculation_status": status,
        "calculation_detail": detail,
    }


def validate_extraction(
    matrix_rows: list[dict[str, Any]],
    partner_detail_rows: list[dict[str, Any]],
    year_sheets: list[str],
    partner_detail_sheets: list[str],
) -> list[str]:
    warnings: list[str] = []
    if not year_sheets:
        warnings.append("No year COGS sheets detected")
    if not partner_detail_sheets:
        warnings.append("No partner detail sheets detected")
    if not matrix_rows:
        warnings.append("No year matrix rows parsed")
    if not partner_detail_rows:
        warnings.append("No partner detail rows parsed")
    return warnings


def normalize_date_value(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return normalize_text(value)


def year_from_month(value: datetime | date) -> int:
    return value.year


def month_from_value(value: datetime | date) -> int:
    return value.month


def month_name(month_num: int) -> str:
    return date(2000, month_num, 1).strftime("%B")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_text(value).lower()).strip()


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("$", "")
        if text in {"", "-", "N/A", "#N/A"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(DivisionCOGSConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
