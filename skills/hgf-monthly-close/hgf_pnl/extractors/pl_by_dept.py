from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


MONTH_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|"
    r"nov(?:ember)?|dec(?:ember)?"
    r")\b",
    re.IGNORECASE,
)


class PLByDeptConfig(BaseModel):
    """Agent-adjustable extraction rules for department matrix P&L workbooks."""

    sheet_name: str | None = Field(default=None, description="Exact sheet name to parse.")
    sheet_name_keywords: list[str] = Field(
        default_factory=lambda: ["profit", "loss", "department"],
        description="Fallback keywords used to pick a sheet when sheet_name is unset.",
    )
    header_keywords: list[str] = Field(
        default_factory=lambda: ["total", "dept", "department", "online", "corp", "production"],
        description="Terms that make a candidate header row more likely.",
    )
    line_item_keywords: list[str] = Field(
        default_factory=lambda: ["sales", "income", "cost", "goods", "profit", "expense"],
        description="Terms that make a column more likely to be the line-item column.",
    )
    max_header_scan_rows: int = 25
    min_department_columns: int = 2
    include_total_columns: bool = True
    total_column_patterns: list[str] = Field(default_factory=lambda: [r"^total(?:\b|$)"])
    skip_line_patterns: list[str] = Field(
        default_factory=lambda: [
            r"^$",
            r"^\s*thursday,",
            r"^\s*monday,",
            r"^\s*tuesday,",
            r"^\s*wednesday,",
            r"^\s*friday,",
            r"^\s*saturday,",
            r"^\s*sunday,",
            r"^\s*accrual basis",
        ],
        description="Line labels to skip entirely.",
    )
    section_patterns: list[str] = Field(
        default_factory=lambda: [
            r"^income$",
            r"^cost of goods sold$",
            r"^expenses?$",
            r"^other income$",
            r"^other expenses?$",
        ],
        description="Rows matching these labels are section headings.",
    )
    stop_after_blank_rows: int = 8
    preserve_zero_amounts: bool = True
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "PLByDeptConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass(frozen=True)
class ExtractedCell:
    value: Any
    coordinate: str
    formula: str | None = None
    fill_color: str | None = None


@dataclass
class PLByDeptExtraction:
    path: Path
    sheet_name: str
    report_title: str | None
    report_period: str | None
    header_row: int
    line_item_column: int
    department_columns: dict[int, str]
    rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    def to_polars(self) -> pl.DataFrame:
        return pl.DataFrame(self.rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheet_name": self.sheet_name,
            "report_title": self.report_title,
            "report_period": self.report_period,
            "header_row": self.header_row,
            "line_item_column": self.line_item_column,
            "department_columns": self.department_columns,
            "warnings": self.warnings,
            "rows": self.rows,
        }


def extract_pl_by_dept(path: Path, config: PLByDeptConfig | None = None) -> PLByDeptExtraction:
    config = config or PLByDeptConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        sheet_name = choose_sheet(workbook_values.sheetnames, config)
        ws_values = workbook_values[sheet_name]
        ws_formulas = workbook_formulas[sheet_name]
        header_row, line_item_col, department_cols = detect_layout(ws_values, config)
        warnings = validate_layout(ws_values, header_row, line_item_col, department_cols, config)
        title, period = detect_report_title_period(ws_values, header_row)
        rows = extract_rows(
            source_path=path,
            ws_values=ws_values,
            ws_formulas=ws_formulas,
            header_row=header_row,
            line_item_col=line_item_col,
            department_cols=department_cols,
            config=config,
            formula_evaluator=formula_evaluator,
        )
        return PLByDeptExtraction(
            path=path,
            sheet_name=sheet_name,
            report_title=title,
            report_period=period,
            header_row=header_row,
            line_item_column=line_item_col,
            department_columns=department_cols,
            rows=rows,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def choose_sheet(sheet_names: list[str], config: PLByDeptConfig) -> str:
    if config.sheet_name:
        if config.sheet_name not in sheet_names:
            raise ValueError(f"Sheet {config.sheet_name!r} not found. Available sheets: {sheet_names}")
        return config.sheet_name

    lowered_keywords = [keyword.lower() for keyword in config.sheet_name_keywords]
    scored: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        lowered = sheet_name.lower()
        score = sum(keyword in lowered for keyword in lowered_keywords)
        scored.append((score, sheet_name))
    scored.sort(reverse=True)
    if scored and scored[0][0] > 0:
        return scored[0][1]
    return sheet_names[0]


def detect_layout(
    ws: Worksheet, config: PLByDeptConfig
) -> tuple[int, int, dict[int, str]]:
    header_row = detect_header_row(ws, config)
    line_item_col = detect_line_item_column(ws, header_row, config)
    department_cols = detect_department_columns(ws, header_row, line_item_col, config)
    if len(department_cols) < config.min_department_columns:
        raise ValueError(
            f"Detected only {len(department_cols)} department columns on row {header_row}; "
            f"expected at least {config.min_department_columns}"
        )
    return header_row, line_item_col, department_cols


def detect_header_row(ws: Worksheet, config: PLByDeptConfig) -> int:
    candidates: list[tuple[float, int]] = []
    for row_idx in range(1, min(ws.max_row, config.max_header_scan_rows) + 1):
        cells = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        non_empty = [normalize_text(value) for value in cells if normalize_text(value)]
        if len(non_empty) < config.min_department_columns:
            continue

        numeric_count = sum(is_number(value) for value in cells)
        text_count = len(non_empty) - numeric_count
        keyword_score = sum(
            keyword.lower() in " ".join(non_empty).lower() for keyword in config.header_keywords
        )
        first_cell_blank_bonus = 2 if not normalize_text(cells[0]) else 0
        dense_text_bonus = min(text_count, 12)
        score = dense_text_bonus + (keyword_score * 3) + first_cell_blank_bonus - numeric_count
        candidates.append((score, row_idx))

    if not candidates:
        raise ValueError("Could not detect a plausible department header row")
    candidates.sort(reverse=True)
    return candidates[0][1]


def detect_line_item_column(ws: Worksheet, header_row: int, config: PLByDeptConfig) -> int:
    search_end = min(ws.max_column, 5)
    scores: list[tuple[float, int]] = []
    for col_idx in range(1, search_end + 1):
        values = [
            normalize_text(ws.cell(row_idx, col_idx).value)
            for row_idx in range(header_row + 1, min(ws.max_row, header_row + 40) + 1)
        ]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        keyword_score = sum(
            any(keyword.lower() in value.lower() for keyword in config.line_item_keywords)
            for value in non_empty
        )
        numeric_penalty = sum(is_number(value) for value in non_empty)
        left_bonus = max(0, 6 - col_idx)
        scores.append((len(non_empty) + (keyword_score * 2) + left_bonus - numeric_penalty, col_idx))

    if not scores:
        return 1
    scores.sort(reverse=True)
    return scores[0][1]


def detect_department_columns(
    ws: Worksheet, header_row: int, line_item_col: int, config: PLByDeptConfig
) -> dict[int, str]:
    department_cols: dict[int, str] = {}
    for col_idx in range(line_item_col + 1, ws.max_column + 1):
        label = normalize_text(ws.cell(header_row, col_idx).value)
        if not label:
            continue
        if not config.include_total_columns and matches_any(label, config.total_column_patterns):
            continue
        numeric_cells_below = 0
        for row_idx in range(header_row + 1, min(ws.max_row, header_row + 60) + 1):
            if is_number(ws.cell(row_idx, col_idx).value):
                numeric_cells_below += 1
        if numeric_cells_below > 0:
            department_cols[col_idx] = label
    return department_cols


def validate_layout(
    ws: Worksheet,
    header_row: int,
    line_item_col: int,
    department_cols: dict[int, str],
    config: PLByDeptConfig,
) -> list[str]:
    warnings: list[str] = []
    if header_row > 12:
        warnings.append(f"Detected header row at {header_row}; that is unusually low in the sheet.")
    if line_item_col != 1:
        warnings.append(f"Detected line-item column {line_item_col}; expected column 1 in most files.")
    if len(department_cols) <= config.min_department_columns:
        warnings.append(f"Only detected {len(department_cols)} department columns.")
    if any(label.lower().startswith("total") for label in department_cols.values()):
        warnings.append("Detected total columns; downstream consumers should avoid double-counting.")
    return warnings


def detect_report_title_period(ws: Worksheet, header_row: int) -> tuple[str | None, str | None]:
    title_parts: list[str] = []
    period: str | None = None
    for row_idx in range(1, max(1, header_row)):
        row_text = " ".join(
            normalize_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, min(ws.max_column, 8) + 1)
            if normalize_text(ws.cell(row_idx, col_idx).value)
        )
        if not row_text:
            continue
        if MONTH_RE.search(row_text) or re.search(r"\b20\d{2}\b", row_text):
            period = row_text
        else:
            title_parts.append(row_text)
    return (" | ".join(title_parts) if title_parts else None, period)


def extract_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    header_row: int,
    line_item_col: int,
    department_cols: dict[int, str],
    config: PLByDeptConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_section: str | None = None
    blank_run = 0

    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        label = normalize_text(ws_values.cell(row_idx, line_item_col).value)
        row_has_any_amount = any(
            is_number(ws_values.cell(row_idx, col_idx).value) for col_idx in department_cols
        )

        if not label and not row_has_any_amount:
            if formula_evaluator is not None and row_has_formula_amounts(
                ws_formulas, row_idx, department_cols
            ):
                row_has_any_amount = True
            else:
                blank_run += 1
                if blank_run >= config.stop_after_blank_rows:
                    break
                continue

        if not label and not row_has_any_amount:
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            continue
        blank_run = 0

        if matches_any(label, config.skip_line_patterns):
            continue

        is_section = bool(label) and matches_any(label, config.section_patterns)
        if is_section:
            current_section = label
            if not row_has_any_amount:
                continue

        if not label and row_has_any_amount:
            label = f"Unlabeled row {row_idx}"

        for col_idx, department in department_cols.items():
            value_cell = ws_values.cell(row_idx, col_idx)
            formula_cell = ws_formulas.cell(row_idx, col_idx)
            cached_amount = value_cell.value
            calculated_amount: Any = None
            calculation_status: str | None = None
            calculation_detail: str | None = None

            if formula_evaluator is not None and formula_cell.data_type == "f":
                calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
                calculated_amount = calculation.value
                calculation_status = calculation.status
                calculation_detail = calculation.detail

            effective_value = cached_amount
            if (
                config.use_calculated_formula_values
                and calculation_status == "ok"
                and calculated_amount is not None
            ):
                effective_value = calculated_amount

            if effective_value is None or effective_value == "":
                continue
            if not config.preserve_zero_amounts and is_zero(effective_value):
                continue
            if not is_number(effective_value):
                continue

            label_cell = ws_values.cell(row_idx, line_item_col)
            rows.append(
                {
                    "source_file": str(source_path),
                    "sheet": ws_values.title,
                    "row": row_idx,
                    "line_item": label,
                    "section": current_section,
                    "department": department,
                    "amount": float(effective_value),
                    "cached_amount": float(cached_amount) if is_number(cached_amount) else None,
                    "calculated_amount": (
                        float(calculated_amount)
                        if is_number(calculated_amount) and not is_formula_sentinel(calculated_amount)
                        else None
                    ),
                    "calculation_status": calculation_status,
                    "calculation_detail": calculation_detail,
                    "is_total_column": matches_any(department, config.total_column_patterns),
                    "is_section_row": is_section,
                    "line_item_cell": label_cell.coordinate,
                    "amount_cell": value_cell.coordinate,
                    "formula": formula_cell.value if formula_cell.data_type == "f" else None,
                    "fill_color": get_fill_color(formula_cell),
                }
            )

    return rows


def row_has_formula_amounts(ws_formulas: Worksheet, row_idx: int, department_cols: dict[int, str]) -> bool:
    return any(ws_formulas.cell(row_idx, col_idx).data_type == "f" for col_idx in department_cols)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def is_number(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return False
        try:
            float(text)
        except ValueError:
            return False
        return True
    return False


def is_zero(value: Any) -> bool:
    if not is_number(value):
        return False
    return float(value) == 0.0


def matches_any(value: str, patterns: list[str]) -> bool:
    return any(re.search(pattern, value, flags=re.IGNORECASE) for pattern in patterns)


def get_fill_color(cell: Cell) -> str | None:
    fill = cell.fill.fgColor
    if fill.type == "rgb" and fill.rgb not in {"00000000", "FFFFFFFF"}:
        return fill.rgb
    return None


def write_default_config(path: Path) -> None:
    path.write_text(
        json.dumps(PLByDeptConfig().model_dump(), indent=2) + "\n",
        encoding="utf-8",
    )
