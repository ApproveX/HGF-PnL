from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any, Literal

import polars as pl
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


class DeclaredTotal(BaseModel):
    """Agent-provided total parsed from natural-language instructions."""

    group_name: str
    amount: float
    amount_column: str = "amount"
    tolerance: float = 1.0


class RowGroupRule(BaseModel):
    """Configurable rule for assigning reviewed GL rows to extraction groups."""

    name: str
    description: str = ""
    match_mode: Literal["any", "all"] = "any"
    fill_colors: list[str] = Field(default_factory=list)
    comment_patterns: list[str] = Field(default_factory=list)
    comment_column: str = "comments"
    nonblank_columns: list[str] = Field(default_factory=list)
    blank_columns: list[str] = Field(default_factory=list)


class AddbacksGLConfig(BaseModel):
    """Agent-adjustable extraction rules for reviewed GL addback workbooks."""

    sheet_name: str | None = None
    sheet_name_keywords: list[str] = Field(
        default_factory=lambda: ["new month", "reviewed", "done", "general ledger"]
    )
    max_header_scan_rows: int = 12
    minimum_header_matches: int = 5
    stop_after_blank_rows: int = 50
    include_blank_rows: bool = False
    header_aliases: dict[str, list[str]] = Field(
        default_factory=lambda: {
            "date": ["date"],
            "transaction_type": ["transaction type", "type"],
            "num": ["num", "number", "no"],
            "name": ["name", "vendor", "customer"],
            "memo_description": ["memo/description", "memo description", "memo", "description"],
            "account": ["split", "account", "gl account"],
            "amount": ["amount", "debit", "credit"],
            "department": ["department", "dept"],
            "expected_account": ["expected account", "account edit", "new account"],
            "expected_department": [
                "expected department",
                "expected dept",
                "department edit",
                "new department",
            ],
            "comments": ["comments", "comment", "notes", "note"],
        }
    )
    row_group_rules: list[RowGroupRule] = Field(
        default_factory=lambda: [
            RowGroupRule(
                name="addbacks",
                description="Rows explicitly marked as addbacks in the accountant review comments.",
                comment_patterns=[r"^addback$"],
                comment_column="comments",
            ),
            RowGroupRule(
                name="red_addback_color_rows",
                description="Rows with red/pink review fill, useful for validating addback instructions.",
                fill_colors=["FFF4CCCC"],
            ),
            RowGroupRule(
                name="unknown_charges",
                description="Rows with magenta review fill.",
                fill_colors=["FFFF00FF"],
            ),
            RowGroupRule(
                name="account_department_edits",
                description="Rows with yellow review fill for expected account/department edits.",
                fill_colors=["FFFFFF00"],
            ),
            RowGroupRule(
                name="other_review_rows",
                description="Rows with blue review fill not described by the March email thread.",
                fill_colors=["FFC9DAF8"],
            ),
        ]
    )
    declared_totals: list[DeclaredTotal] = Field(default_factory=list)
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "AddbacksGLConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class AddbacksGLExtraction:
    path: Path
    sheet_name: str
    header_row: int
    header_map: dict[str, int]
    rows: list[dict[str, Any]]
    group_rows: list[dict[str, Any]]
    group_summaries: list[dict[str, Any]]
    color_summaries: list[dict[str, Any]]
    reconciliations: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    @property
    def ledger(self) -> pl.DataFrame:
        return pl.DataFrame(self.rows, infer_schema_length=None)

    @property
    def groups(self) -> pl.DataFrame:
        return pl.DataFrame(self.group_rows, infer_schema_length=None)

    @property
    def summaries(self) -> pl.DataFrame:
        return pl.DataFrame(self.group_summaries, infer_schema_length=None)

    def group(self, name: str) -> pl.DataFrame:
        if not self.group_rows:
            return pl.DataFrame()
        return self.groups.filter(pl.col("group_name") == name)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "header_map": self.header_map,
            "warnings": self.warnings,
            "group_summaries": self.group_summaries,
            "color_summaries": self.color_summaries,
            "reconciliations": self.reconciliations,
            "rows": self.rows,
            "group_rows": self.group_rows,
        }


def extract_addbacks_gl(
    path: Path,
    config: AddbacksGLConfig | None = None,
) -> AddbacksGLExtraction:
    config = config or AddbacksGLConfig()
    workbook_values = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        sheet_name = choose_sheet(workbook_values, config)
        if sheet_name is None:
            raise ValueError("Could not find reviewed GL sheet")

        ws_values = workbook_values[sheet_name]
        ws_formulas = workbook_formulas[sheet_name]
        header_row, header_map = detect_header_row(ws_values, config)
        rows = extract_ledger_rows(
            path,
            ws_values,
            ws_formulas,
            header_row,
            header_map,
            config,
            formula_evaluator,
        )
        group_rows = assign_group_rows(rows, config.row_group_rules)
        group_summaries = summarize_groups(group_rows)
        color_summaries = summarize_colors(rows)
        reconciliations = reconcile_declared_totals(group_rows, config.declared_totals)
        warnings = validate_extraction(rows, group_rows, group_summaries, reconciliations, header_map)

        return AddbacksGLExtraction(
            path=path,
            sheet_name=sheet_name,
            header_row=header_row,
            header_map=header_map,
            rows=rows,
            group_rows=group_rows,
            group_summaries=group_summaries,
            color_summaries=color_summaries,
            reconciliations=reconciliations,
            warnings=warnings,
        )
    finally:
        workbook_values.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def choose_sheet(workbook: Any, config: AddbacksGLConfig) -> str | None:
    if config.sheet_name:
        return config.sheet_name if config.sheet_name in workbook.sheetnames else None

    scored: list[tuple[int, int, str]] = []
    for sheet_name in workbook.sheetnames:
        normalized = normalize_key(sheet_name)
        keyword_score = sum(
            100 for keyword in config.sheet_name_keywords if normalize_key(keyword) in normalized
        )
        colored_score = min(count_colored_cells(workbook[sheet_name]), 50)
        scored.append((keyword_score + colored_score, colored_score, sheet_name))
    scored.sort(reverse=True)
    if not scored:
        return None
    return scored[0][2]


def count_colored_cells(ws: Worksheet) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if normalized_fill_color(cell):
                count += 1
    return count


def detect_header_row(ws: Worksheet, config: AddbacksGLConfig) -> tuple[int, dict[str, int]]:
    candidates: list[tuple[int, int, dict[str, int]]] = []
    max_row = min(ws.max_row, config.max_header_scan_rows)
    for row_idx in range(1, max_row + 1):
        header_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            raw_header = normalize_text(ws.cell(row_idx, col_idx).value)
            if not raw_header:
                continue
            canonical = canonical_header(raw_header, config.header_aliases)
            if canonical and canonical not in header_map:
                header_map[canonical] = col_idx
        if len(header_map) >= config.minimum_header_matches:
            candidates.append((len(header_map), row_idx, header_map))

    if not candidates:
        raise ValueError(f"Could not detect reviewed GL header row in {ws.title!r}")
    candidates.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    _, row_idx, header_map = candidates[0]
    return row_idx, header_map


def canonical_header(raw_header: str, aliases: dict[str, list[str]]) -> str | None:
    normalized = normalize_key(raw_header)
    for canonical, candidates in aliases.items():
        for candidate in candidates:
            candidate_key = normalize_key(candidate)
            if normalized == candidate_key:
                return canonical
    return None


def extract_ledger_rows(
    path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    config: AddbacksGLConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_run = 0
    current_account_section = ""
    account_section_col = min(header_map.values()) - 1 if header_map and min(header_map.values()) > 1 else None
    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        row_values = {
            canonical: ws_values.cell(row_idx, col_idx).value
            for canonical, col_idx in header_map.items()
        }
        colors = row_fill_colors(ws_values, row_idx)
        has_data = any(value not in (None, "") for value in row_values.values())
        has_review_color = bool(colors)
        if not has_data and not has_review_color:
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            if not config.include_blank_rows:
                continue
        else:
            blank_run = 0

        section_label = ""
        if account_section_col is not None:
            section_label = clean_account_section(ws_values.cell(row_idx, account_section_col).value)
        if not section_label:
            section_label = account_section_label(row_values)
        if section_label:
            current_account_section = section_label
        is_section_row = bool(section_label) and not normalize_text(row_values.get("transaction_type"))

        amount_info = formula_cell_info(
            ws_values,
            ws_formulas,
            row_idx,
            header_map.get("amount"),
            formula_evaluator,
            config,
        )
        amount = to_float(amount_info["amount"])
        row = {
            "source_file": str(path),
            "sheet": ws_values.title,
            "row": row_idx,
            "source_row": row_idx,
            "account_section": current_account_section,
            "is_account_section_row": is_section_row,
            "dominant_fill_color": dominant_fill_color(colors),
            "row_fill_colors": sorted(colors),
            "colored_cells": color_cell_refs(ws_values, row_idx),
            "amount": amount,
            "raw_amount": normalize_text(amount_info["amount"]),
            "amount_cell": cell_ref(row_idx, header_map.get("amount")),
            "formula": amount_info["formula"],
            "cached_amount": to_float(amount_info["cached_value"]),
            "calculated_amount": to_float(amount_info["calculated_value"]),
            "calculation_status": amount_info["calculation_status"],
            "calculation_detail": amount_info["calculation_detail"],
        }
        for canonical, value in row_values.items():
            if canonical == "amount":
                continue
            row[canonical] = normalize_value(value)
            row[f"{canonical}_cell"] = cell_ref(row_idx, header_map[canonical])
        rows.append(row)
    return rows


def account_section_label(row_values: dict[str, Any]) -> str:
    date_value = row_values.get("date")
    label = normalize_text(date_value)
    if not label or looks_like_date(label):
        return ""
    detail_columns = [
        "transaction_type",
        "num",
        "name",
        "memo_description",
        "account",
        "department",
        "expected_account",
        "expected_department",
        "comments",
    ]
    if any(normalize_text(row_values.get(column)) for column in detail_columns):
        return ""
    return label


def clean_account_section(value: Any) -> str:
    label = normalize_text(value)
    if not label:
        return ""
    normalized = normalize_key(label)
    if normalized.startswith("total for"):
        return ""
    if normalized in {"total", "beginning balance"}:
        return ""
    return label


def formula_cell_info(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    col_idx: int | None,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    config: AddbacksGLConfig,
) -> dict[str, Any]:
    if col_idx is None:
        return {
            "amount": None,
            "formula": None,
            "cached_value": None,
            "calculated_value": None,
            "calculation_status": "missing_amount_column",
            "calculation_detail": None,
        }
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    calculated_value = None
    status = "not_formula"
    detail = None
    amount = cached_value

    if formula_evaluator is not None and formula_cell.data_type == "f":
        calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
        calculated_value = calculation.value
        status = calculation.status
        detail = calculation.detail
        if (
            config.use_calculated_formula_values
            and calculation.status == "ok"
            and not is_formula_sentinel(calculation.value)
        ):
            amount = calculation.value

    return {
        "amount": amount,
        "formula": formula,
        "cached_value": cached_value,
        "calculated_value": calculated_value,
        "calculation_status": status,
        "calculation_detail": detail,
    }


def assign_group_rows(
    rows: list[dict[str, Any]],
    rules: list[RowGroupRule],
) -> list[dict[str, Any]]:
    group_rows: list[dict[str, Any]] = []
    for row in rows:
        for rule in rules:
            if row_matches_rule(row, rule):
                group_row = dict(row)
                group_row["group_name"] = rule.name
                group_row["group_description"] = rule.description
                group_rows.append(group_row)
    return group_rows


def row_matches_rule(row: dict[str, Any], rule: RowGroupRule) -> bool:
    checks: list[bool] = []
    configured = False

    if rule.fill_colors:
        configured = True
        row_colors = {normalize_color(color) for color in row.get("row_fill_colors", [])}
        rule_colors = {normalize_color(color) for color in rule.fill_colors}
        checks.append(bool(row_colors & rule_colors))

    if rule.comment_patterns:
        configured = True
        comment = normalize_text(row.get(rule.comment_column))
        checks.append(
            any(re.search(pattern, comment, flags=re.IGNORECASE) for pattern in rule.comment_patterns)
        )

    for column in rule.nonblank_columns:
        configured = True
        checks.append(bool(normalize_text(row.get(column))))

    for column in rule.blank_columns:
        configured = True
        checks.append(not bool(normalize_text(row.get(column))))

    if not configured:
        return False
    return all(checks) if rule.match_mode == "all" else any(checks)


def summarize_groups(group_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for row in group_rows:
        group_name = row["group_name"]
        item = summary.setdefault(
            group_name,
            {
                "group_name": group_name,
                "description": row.get("group_description", ""),
                "row_count": 0,
                "amount_total": 0.0,
                "amount_non_null_count": 0,
            },
        )
        item["row_count"] += 1
        amount = row.get("amount")
        if amount is not None:
            item["amount_total"] += float(amount)
            item["amount_non_null_count"] += 1
    return list(summary.values())


def summarize_colors(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        for color in row.get("row_fill_colors", []):
            item = summary.setdefault(
                color,
                {
                    "fill_color": color,
                    "row_count": 0,
                    "amount_total": 0.0,
                    "amount_non_null_count": 0,
                },
            )
            item["row_count"] += 1
            amount = row.get("amount")
            if amount is not None:
                item["amount_total"] += float(amount)
                item["amount_non_null_count"] += 1
    return sorted(summary.values(), key=lambda item: item["row_count"], reverse=True)


def reconcile_declared_totals(
    group_rows: list[dict[str, Any]],
    declared_totals: list[DeclaredTotal],
) -> list[dict[str, Any]]:
    reconciliations: list[dict[str, Any]] = []
    for declared in declared_totals:
        matched = [
            row
            for row in group_rows
            if row.get("group_name") == declared.group_name
            and row.get(declared.amount_column) is not None
        ]
        extracted_total = sum(float(row[declared.amount_column]) for row in matched)
        difference = extracted_total - declared.amount
        reconciliations.append(
            {
                "group_name": declared.group_name,
                "amount_column": declared.amount_column,
                "declared_total": declared.amount,
                "extracted_total": extracted_total,
                "difference": difference,
                "tolerance": declared.tolerance,
                "status": "ok" if abs(difference) <= declared.tolerance else "mismatch",
                "row_count": len(matched),
            }
        )
    return reconciliations


def validate_extraction(
    rows: list[dict[str, Any]],
    group_rows: list[dict[str, Any]],
    group_summaries: list[dict[str, Any]],
    reconciliations: list[dict[str, Any]],
    header_map: dict[str, int],
) -> list[str]:
    warnings: list[str] = []
    if not rows:
        warnings.append("No reviewed GL rows parsed")
    if not group_rows:
        warnings.append("No row groups matched")
    if "amount" not in header_map:
        warnings.append("No amount column detected")
    if "comments" not in header_map:
        warnings.append("No comments column detected; comment-based addback extraction is unavailable")

    summaries = {summary["group_name"]: summary for summary in group_summaries}
    addbacks = summaries.get("addbacks")
    red_rows = summaries.get("red_addback_color_rows")
    if addbacks and red_rows:
        difference = red_rows["amount_total"] - addbacks["amount_total"]
        if abs(difference) > 1:
            warnings.append(
                "Red/pink row total differs from comment-based addback total by "
                f"{difference:,.2f}"
            )

    for reconciliation in reconciliations:
        if reconciliation["status"] != "ok":
            warnings.append(
                f"{reconciliation['group_name']} declared total mismatch: "
                f"extracted {reconciliation['extracted_total']:,.2f}, "
                f"declared {reconciliation['declared_total']:,.2f}"
            )
    return warnings


def row_fill_colors(ws: Worksheet, row_idx: int) -> set[str]:
    colors: set[str] = set()
    for col_idx in range(1, ws.max_column + 1):
        color = normalized_fill_color(ws.cell(row_idx, col_idx))
        if color:
            colors.add(color)
    return colors


def color_cell_refs(ws: Worksheet, row_idx: int) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    for col_idx in range(1, ws.max_column + 1):
        color = normalized_fill_color(ws.cell(row_idx, col_idx))
        if color:
            refs.append({"cell": cell_ref(row_idx, col_idx), "fill_color": color})
    return refs


def dominant_fill_color(colors: set[str]) -> str | None:
    if not colors:
        return None
    return sorted(colors)[0]


def normalized_fill_color(cell: Cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        normalized = normalize_color(color.rgb)
        if normalized in {"00000000", "FFFFFFFF"}:
            return None
        return normalized
    if color.type == "indexed" and color.indexed is not None:
        return f"INDEXED:{color.indexed}"
    if color.type == "theme" and color.theme is not None:
        return f"THEME:{color.theme}:{color.tint or 0}"
    return None


def cell_ref(row_idx: int, col_idx: int | None) -> str | None:
    if col_idx is None:
        return None
    return f"{get_column_letter(col_idx)}{row_idx}"


def normalize_color(value: Any) -> str:
    return normalize_text(value).replace("#", "").upper()


def normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, int | float):
        return value
    return normalize_text(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_text(value).lower()).strip()


def looks_like_date(value: str) -> bool:
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", value):
        return True
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:T.*)?", value):
        return True
    return False


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("$", "")
        if text in {"", "-", "N/A", "#N/A"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(AddbacksGLConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
