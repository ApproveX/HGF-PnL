from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import json
import re
from pathlib import Path
from typing import Any, Literal

import polars as pl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from rapidfuzz import fuzz

from hgf_pnl.formulas import WorkbookFormulaEvaluator, is_formula_sentinel


SheetRole = Literal["summary", "details", "usa_stock"]


ROLE_KEYWORDS: dict[SheetRole, list[str]] = {
    "summary": ["summary"],
    "details": ["detail", "details"],
    "usa_stock": ["usa", "stock"],
}


CANONICAL_ALIASES: dict[str, list[str]] = {
    "internal_po": ["internal po", "po", "internal po #"],
    "account": ["account", "customer", "customer account"],
    "revenue": ["revenue", "sales", "gross sales"],
    "production_cost": ["production cost", "prod cost", "product cost"],
    "shipping_cost": ["shipping cost", "shipping cost ", "shipping"],
    "tariff": ["tariff", "tariffs"],
    "total_cost": ["total cost", "total costs"],
    "gross_margin_pct": ["gm %", "gross margin %", "gross margin percentage", "gross margin"],
    "gross_margin_amount": ["gm $", "gross margin $", "gross margin dollars"],
}


ROLE_REQUIRED_COLUMNS: dict[SheetRole, set[str]] = {
    "summary": {"account", "revenue", "production_cost", "total_cost"},
    "details": {"internal_po", "account", "revenue", "production_cost", "total_cost"},
    "usa_stock": {"internal_po", "account", "revenue", "production_cost", "total_cost"},
}


class THRevenueConfig(BaseModel):
    """Agent-adjustable extraction rules for Trend House revenue reports."""

    sheet_roles: dict[SheetRole, list[str]] = Field(default_factory=lambda: ROLE_KEYWORDS.copy())
    column_aliases: dict[str, list[str]] = Field(default_factory=lambda: CANONICAL_ALIASES.copy())
    max_header_scan_rows: int = 12
    fuzzy_header_threshold: int = 86
    include_total_rows: bool = True
    total_row_patterns: list[str] = Field(default_factory=lambda: [r"^total\b", r"^grand total\b"])
    stop_after_blank_rows: int = 5
    preserve_source_columns: bool = True
    calculate_formulas: bool = True
    use_calculated_formula_values: bool = True

    @classmethod
    def from_json_file(cls, path: Path | None) -> "THRevenueConfig":
        if path is None:
            return cls()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class SheetExtraction:
    role: SheetRole
    sheet_name: str
    header_row: int
    column_map: dict[str, int]
    source_headers: dict[str, str]
    rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    def to_polars(self) -> pl.DataFrame:
        return pl.DataFrame(self.rows)


@dataclass
class THRevenueExtraction:
    path: Path
    sheets: dict[SheetRole, SheetExtraction]
    warnings: list[str] = field(default_factory=list)

    @property
    def account_summary(self) -> pl.DataFrame:
        return self.sheets["summary"].to_polars()

    @property
    def po_details(self) -> pl.DataFrame:
        if "details" not in self.sheets:
            return pl.DataFrame()
        return self.sheets["details"].to_polars()

    @property
    def usa_stock(self) -> pl.DataFrame:
        if "usa_stock" not in self.sheets:
            return pl.DataFrame()
        return self.sheets["usa_stock"].to_polars()

    def all_rows(self) -> pl.DataFrame:
        frames = [sheet.to_polars() for sheet in self.sheets.values() if sheet.rows]
        if not frames:
            return pl.DataFrame()
        return pl.concat(frames, how="diagonal_relaxed")

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "warnings": self.warnings,
            "sheets": {
                role: {
                    "sheet_name": sheet.sheet_name,
                    "header_row": sheet.header_row,
                    "column_map": sheet.column_map,
                    "source_headers": sheet.source_headers,
                    "warnings": sheet.warnings,
                    "rows": sheet.rows,
                }
                for role, sheet in self.sheets.items()
            },
        }


def extract_th_revenue(
    path: Path,
    config: THRevenueConfig | None = None,
) -> THRevenueExtraction:
    config = config or THRevenueConfig()
    workbook = load_workbook(path, read_only=False, data_only=True)
    workbook_formulas = load_workbook(path, read_only=False, data_only=False)
    formula_evaluator = WorkbookFormulaEvaluator(workbook_formulas) if config.calculate_formulas else None
    try:
        role_to_sheet = choose_role_sheets(workbook.sheetnames, config)
        warnings: list[str] = []
        sheets: dict[SheetRole, SheetExtraction] = {}
        for role, sheet_name in role_to_sheet.items():
            sheet = extract_sheet(
                path=path,
                ws_values=workbook[sheet_name],
                ws_formulas=workbook_formulas[sheet_name],
                role=role,
                config=config,
                formula_evaluator=formula_evaluator,
            )
            sheets[role] = sheet
            warnings.extend(f"{role}: {warning}" for warning in sheet.warnings)

        missing_roles = set(ROLE_KEYWORDS) - set(sheets)
        if missing_roles:
            warnings.append(f"Missing expected sheets for roles: {sorted(missing_roles)}")

        return THRevenueExtraction(path=path, sheets=sheets, warnings=warnings)
    finally:
        workbook.close()
        if formula_evaluator is not None:
            formula_evaluator.close()
        workbook_formulas.close()


def choose_role_sheets(sheet_names: list[str], config: THRevenueConfig) -> dict[SheetRole, str]:
    chosen: dict[SheetRole, str] = {}
    used: set[str] = set()
    for role, keywords in config.sheet_roles.items():
        scored: list[tuple[int, str]] = []
        lowered_keywords = [normalize_key(keyword) for keyword in keywords]
        for sheet_name in sheet_names:
            if sheet_name in used:
                continue
            normalized = normalize_key(sheet_name)
            score = sum(keyword in normalized for keyword in lowered_keywords)
            fuzzy_score = max(
                (fuzz.partial_ratio(keyword, normalized) for keyword in lowered_keywords),
                default=0,
            )
            scored.append((score * 100 + int(fuzzy_score), sheet_name))
        scored.sort(reverse=True)
        if scored and scored[0][0] >= config.fuzzy_header_threshold:
            chosen[role] = scored[0][1]
            used.add(scored[0][1])
    return chosen


def extract_sheet(
    path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    role: SheetRole,
    config: THRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None = None,
) -> SheetExtraction:
    header_row, raw_headers = detect_header_row(ws_values, role, config)
    column_map, source_headers = map_headers(raw_headers, config)
    warnings = validate_columns(role, column_map)
    rows = extract_table_rows(
        source_path=path,
        ws_values=ws_values,
        ws_formulas=ws_formulas,
        role=role,
        header_row=header_row,
        column_map=column_map,
        source_headers=source_headers,
        config=config,
        formula_evaluator=formula_evaluator,
    )
    return SheetExtraction(
        role=role,
        sheet_name=ws_values.title,
        header_row=header_row,
        column_map=column_map,
        source_headers=source_headers,
        rows=rows,
        warnings=warnings,
    )


def detect_header_row(
    ws: Worksheet,
    role: SheetRole,
    config: THRevenueConfig,
) -> tuple[int, dict[int, str]]:
    candidates: list[tuple[int, int, dict[int, str]]] = []
    for row_idx in range(1, min(ws.max_row, config.max_header_scan_rows) + 1):
        headers = {
            col_idx: normalize_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
            if normalize_text(ws.cell(row_idx, col_idx).value)
        }
        if not headers:
            continue
        mapped, _ = map_headers(headers, config)
        required_hits = len(ROLE_REQUIRED_COLUMNS[role] & set(mapped))
        alias_hits = len(mapped)
        text_density = min(len(headers), 12)
        numeric_penalty = sum(is_number(header) for header in headers.values())
        score = required_hits * 20 + alias_hits * 5 + text_density - numeric_penalty
        candidates.append((score, row_idx, headers))

    if not candidates:
        raise ValueError(f"Could not detect header row for sheet {ws.title!r}")
    candidates.sort(reverse=True)
    return candidates[0][1], candidates[0][2]


def map_headers(
    headers: dict[int, str],
    config: THRevenueConfig,
) -> tuple[dict[str, int], dict[str, str]]:
    mapped: dict[str, int] = {}
    source_headers: dict[str, str] = {}
    for col_idx, header in headers.items():
        canonical = match_header(header, config.column_aliases, config.fuzzy_header_threshold)
        if canonical and canonical not in mapped:
            mapped[canonical] = col_idx
            source_headers[canonical] = header
    return mapped, source_headers


def match_header(
    header: str,
    aliases: dict[str, list[str]],
    threshold: int,
) -> str | None:
    normalized_header = normalize_key(header)
    best: tuple[int, str] | None = None
    for canonical, candidates in aliases.items():
        for candidate in candidates:
            normalized_candidate = normalize_key(candidate)
            if normalized_header == normalized_candidate:
                return canonical
            if normalized_candidate in normalized_header or normalized_header in normalized_candidate:
                score = 95
            else:
                score = int(fuzz.token_sort_ratio(normalized_header, normalized_candidate))
            if best is None or score > best[0]:
                best = (score, canonical)
    if best and best[0] >= threshold:
        return best[1]
    return None


def validate_columns(role: SheetRole, column_map: dict[str, int]) -> list[str]:
    missing = ROLE_REQUIRED_COLUMNS[role] - set(column_map)
    if missing:
        return [f"Missing expected columns: {sorted(missing)}"]
    return []


def extract_table_rows(
    source_path: Path,
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    role: SheetRole,
    header_row: int,
    column_map: dict[str, int],
    source_headers: dict[str, str],
    config: THRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_run = 0
    numeric_fields = {
        "revenue",
        "production_cost",
        "shipping_cost",
        "tariff",
        "total_cost",
        "gross_margin_pct",
        "gross_margin_amount",
    }
    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        row_values = {
            canonical: ws_values.cell(row_idx, col_idx).value for canonical, col_idx in column_map.items()
        }
        if all(value in (None, "") for value in row_values.values()):
            blank_run += 1
            if blank_run >= config.stop_after_blank_rows:
                break
            continue
        blank_run = 0

        account = normalize_text(row_values.get("account"))
        internal_po = normalize_text(row_values.get("internal_po"))
        if not account and not internal_po:
            continue

        is_total_row = matches_any(account, config.total_row_patterns) or matches_any(
            internal_po, config.total_row_patterns
        )
        if is_total_row and not config.include_total_rows:
            continue

        formula_status: dict[str, str] = {}
        formula_detail: dict[str, str | None] = {}
        numeric_values = {
            field: effective_cell_value(
                ws_values=ws_values,
                ws_formulas=ws_formulas,
                row_idx=row_idx,
                column_map=column_map,
                field=field,
                config=config,
                formula_evaluator=formula_evaluator,
                formula_status=formula_status,
                formula_detail=formula_detail,
            )
            for field in numeric_fields
        }

        record: dict[str, Any] = {
            "source_file": str(source_path),
            "sheet": ws_values.title,
            "role": role,
            "row": row_idx,
            "is_total_row": is_total_row,
            "internal_po": internal_po or None,
            "account": account or None,
            "revenue": to_float(numeric_values.get("revenue")),
            "production_cost": to_float(numeric_values.get("production_cost")),
            "shipping_cost": to_float(numeric_values.get("shipping_cost")),
            "tariff": to_float(numeric_values.get("tariff")),
            "total_cost": to_float(numeric_values.get("total_cost")),
            "gross_margin_pct": to_float(numeric_values.get("gross_margin_pct")),
            "gross_margin_amount": to_float(numeric_values.get("gross_margin_amount")),
            "formula_status": json.dumps(formula_status, sort_keys=True),
            "formula_detail": json.dumps(formula_detail, sort_keys=True),
        }
        record["computed_gross_margin_amount"] = compute_margin_amount(record)
        record["computed_gross_margin_pct"] = compute_margin_pct(record)
        record["validation_warnings"] = json.dumps(validate_row(record))
        if config.preserve_source_columns:
            record["source_headers"] = json.dumps(source_headers, sort_keys=True)
        rows.append(record)
    return rows


def effective_cell_value(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row_idx: int,
    column_map: dict[str, int],
    field: str,
    config: THRevenueConfig,
    formula_evaluator: WorkbookFormulaEvaluator | None,
    formula_status: dict[str, str],
    formula_detail: dict[str, str | None],
) -> Any:
    col_idx = column_map.get(field)
    if col_idx is None:
        return None
    cached_value = ws_values.cell(row_idx, col_idx).value
    formula_cell = ws_formulas.cell(row_idx, col_idx)
    if formula_evaluator is None or formula_cell.data_type != "f":
        return cached_value

    calculation = formula_evaluator.evaluate_cell(ws_formulas.title, formula_cell.coordinate)
    formula_status[field] = calculation.status
    formula_detail[field] = calculation.detail
    if (
        config.use_calculated_formula_values
        and calculation.status == "ok"
        and not is_formula_sentinel(calculation.value)
    ):
        return calculation.value
    return cached_value


def compute_margin_amount(record: dict[str, Any]) -> float | None:
    revenue = record.get("revenue")
    total_cost = record.get("total_cost")
    if revenue is None or total_cost is None:
        return None
    return revenue - total_cost


def compute_margin_pct(record: dict[str, Any]) -> float | None:
    revenue = record.get("revenue")
    margin = compute_margin_amount(record)
    if revenue in (None, 0) or margin is None:
        return None
    return margin / revenue


def validate_row(record: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    margin_amount = record.get("gross_margin_amount")
    computed_margin_amount = record.get("computed_gross_margin_amount")
    if margin_amount is not None and computed_margin_amount is not None:
        if abs(margin_amount - computed_margin_amount) > 0.02:
            warnings.append("gross_margin_amount_mismatch")

    margin_pct = record.get("gross_margin_pct")
    computed_margin_pct = record.get("computed_gross_margin_pct")
    if margin_pct is not None and computed_margin_pct is not None:
        if abs(margin_pct - computed_margin_pct) > 0.0001:
            warnings.append("gross_margin_pct_mismatch")
    return warnings


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def is_number(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return False
        try:
            float(text)
        except ValueError:
            return False
        return True
    return False


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def matches_any(value: str | None, patterns: list[str]) -> bool:
    if not value:
        return False
    return any(re.search(pattern, value, flags=re.IGNORECASE) for pattern in patterns)


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(THRevenueConfig().model_dump(), indent=2) + "\n", encoding="utf-8")
