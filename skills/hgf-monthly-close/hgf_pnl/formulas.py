from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook


FormulaStatus = Literal["ok", "unsupported", "error", "cycle"]


@dataclass(frozen=True)
class FormulaSentinel:
    status: FormulaStatus
    reason: str
    formula: str | None = None
    detail: str | None = None

    def __bool__(self) -> bool:
        return False


UNSUPPORTED_FORMULA = FormulaSentinel("unsupported", "unsupported_formula")


@dataclass(frozen=True)
class FormulaResult:
    value: Any
    status: FormulaStatus
    formula: str | None = None
    detail: str | None = None

    @property
    def ok(self) -> bool:
        return self.status == "ok"


@dataclass(frozen=True)
class Token:
    kind: str
    value: str
    position: int


class FormulaParseError(ValueError):
    pass


class WorkbookFormulaEvaluator:
    """Evaluate a conservative subset of Excel formulas in a same-workbook context."""

    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self._cache: dict[tuple[str, str], FormulaResult] = {}
        self._stack: set[tuple[str, str]] = set()

    @classmethod
    def from_path(cls, path: Path) -> "WorkbookFormulaEvaluator":
        workbook = load_workbook(path, read_only=False, data_only=False)
        return cls(workbook)

    def close(self) -> None:
        self.workbook.close()

    def evaluate_cell(self, sheet_name: str, coordinate: str) -> FormulaResult:
        key = (sheet_name, normalize_coordinate(coordinate))
        if key in self._cache:
            return self._cache[key]
        if key in self._stack:
            return FormulaResult(
                FormulaSentinel("cycle", "formula_cycle", detail=f"{sheet_name}!{coordinate}"),
                "cycle",
                detail=f"{sheet_name}!{coordinate}",
            )

        self._stack.add(key)
        try:
            cell = self.workbook[sheet_name][coordinate]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                result = self.evaluate_formula(cell.value, sheet_name)
            else:
                result = FormulaResult(cell.value, "ok")
            self._cache[key] = result
            return result
        finally:
            self._stack.remove(key)

    def evaluate_formula(self, formula: str, current_sheet: str) -> FormulaResult:
        try:
            parser = FormulaParser(formula, self, current_sheet)
            value = parser.parse()
        except FormulaParseError as exc:
            return FormulaResult(
                FormulaSentinel("unsupported", "parse_error", formula=formula, detail=str(exc)),
                "unsupported",
                formula=formula,
                detail=str(exc),
            )
        except ZeroDivisionError:
            return FormulaResult(
                FormulaSentinel("error", "division_by_zero", formula=formula),
                "error",
                formula=formula,
                detail="division_by_zero",
            )
        except Exception as exc:
            return FormulaResult(
                FormulaSentinel(
                    "error",
                    "evaluation_error",
                    formula=formula,
                    detail=f"{type(exc).__name__}: {exc}",
                ),
                "error",
                formula=formula,
                detail=f"{type(exc).__name__}: {exc}",
            )

        if isinstance(value, FormulaSentinel):
            return FormulaResult(value, value.status, formula=formula, detail=value.detail)
        return FormulaResult(value, "ok", formula=formula)

    def get_reference_value(self, sheet_name: str, coordinate: str) -> Any:
        result = self.evaluate_cell(sheet_name, coordinate)
        return result.value

    def get_range_values(self, sheet_name: str, ref_range: str) -> list[Any]:
        min_col, min_row, max_col, max_row = range_boundaries(strip_dollars(ref_range))
        values: list[Any] = []
        ws = self.workbook[sheet_name]
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                values.append(self.get_reference_value(sheet_name, cell.coordinate))
        return values


class FormulaParser:
    def __init__(
        self,
        formula: str,
        evaluator: WorkbookFormulaEvaluator,
        current_sheet: str,
    ):
        self.formula = formula
        self.evaluator = evaluator
        self.current_sheet = current_sheet
        self.tokens = tokenize(formula)
        self.index = 0

    def parse(self) -> Any:
        if self.match("OP", "="):
            pass
        value = self.parse_expression()
        if not self.at_end():
            token = self.peek()
            raise FormulaParseError(f"Unexpected token {token.value!r} at {token.position}")
        return value

    def parse_expression(self) -> Any:
        return self.parse_additive()

    def parse_additive(self) -> Any:
        value = self.parse_multiplicative()
        while self.match("OP", "+") or self.match("OP", "-"):
            operator = self.previous().value
            right = self.parse_multiplicative()
            value = apply_binary(operator, value, right, self.formula)
        return value

    def parse_multiplicative(self) -> Any:
        value = self.parse_power()
        while self.match("OP", "*") or self.match("OP", "/"):
            operator = self.previous().value
            right = self.parse_power()
            value = apply_binary(operator, value, right, self.formula)
        return value

    def parse_power(self) -> Any:
        value = self.parse_unary()
        while self.match("OP", "^"):
            right = self.parse_unary()
            value = apply_binary("^", value, right, self.formula)
        return value

    def parse_unary(self) -> Any:
        if self.match("OP", "+"):
            return self.parse_unary()
        if self.match("OP", "-"):
            return -to_number(self.parse_unary(), self.formula)
        return self.parse_percent()

    def parse_percent(self) -> Any:
        value = self.parse_primary()
        while self.match("OP", "%"):
            value = to_number(value, self.formula) / 100
        return value

    def parse_primary(self) -> Any:
        if self.match("NUMBER"):
            return parse_number(self.previous().value)
        if self.match("STRING"):
            return self.previous().value
        if self.match("OP", "("):
            value = self.parse_expression()
            self.consume("OP", ")", "Expected ')' after expression")
            return value

        if self.check("IDENT") and self.check_next("OP", "("):
            function_name = self.advance().value.upper()
            return self.parse_function_call(function_name)

        reference = self.try_parse_reference()
        if reference is not None:
            sheet_name, ref = reference
            if ":" in ref:
                return self.evaluator.get_range_values(sheet_name, ref)
            return self.evaluator.get_reference_value(sheet_name, ref)

        token = self.peek()
        raise FormulaParseError(f"Unsupported token {token.value!r} at {token.position}")

    def parse_function_call(self, function_name: str) -> Any:
        self.consume("OP", "(", "Expected '(' after function name")
        if function_name not in SUPPORTED_FUNCTIONS:
            self.skip_balanced_arguments()
            return FormulaSentinel(
                "unsupported",
                "unsupported_function",
                formula=self.formula,
                detail=function_name,
            )

        args: list[Any] = []
        if not self.check("OP", ")"):
            while True:
                args.append(self.parse_expression())
                if not self.match("OP", ","):
                    break
        self.consume("OP", ")", "Expected ')' after function arguments")
        return evaluate_function(function_name, args, self.formula)

    def skip_balanced_arguments(self) -> None:
        depth = 1
        while not self.at_end() and depth:
            token = self.advance()
            if token.kind == "OP" and token.value == "(":
                depth += 1
            elif token.kind == "OP" and token.value == ")":
                depth -= 1

    def try_parse_reference(self) -> tuple[str, str] | None:
        mark = self.index
        sheet_name = self.current_sheet

        if self.check("WORKBOOK"):
            self.advance()
            if self.check("SHEET") and self.check_next("OP", "!"):
                sheet_name = normalize_sheet_reference_name(self.advance().value)
                self.advance()
            elif self.check("IDENT") and self.check_next("OP", "!"):
                sheet_name = normalize_sheet_reference_name(self.advance().value)
                self.advance()
            else:
                self.index = mark
                return None
        elif self.check("SHEET") and self.check_next("OP", "!"):
            sheet_name = normalize_sheet_reference_name(self.advance().value)
            self.advance()
        elif self.check("IDENT") and self.check_next("OP", "!"):
            sheet_name = normalize_sheet_reference_name(self.advance().value)
            self.advance()

        if not self.match("CELL"):
            self.index = mark
            return None

        start_ref = self.previous().value
        if self.match("OP", ":"):
            self.consume("CELL", None, "Expected cell reference after ':'")
            end_ref = self.previous().value
            return sheet_name, f"{start_ref}:{end_ref}"
        return sheet_name, start_ref

    def consume(self, kind: str, value: str | None, message: str) -> Token:
        if self.check(kind, value):
            return self.advance()
        raise FormulaParseError(message)

    def match(self, kind: str, value: str | None = None) -> bool:
        if not self.check(kind, value):
            return False
        self.advance()
        return True

    def check(self, kind: str, value: str | None = None) -> bool:
        if self.at_end():
            return False
        token = self.peek()
        return token.kind == kind and (value is None or token.value == value)

    def check_next(self, kind: str, value: str | None = None) -> bool:
        if self.index + 1 >= len(self.tokens):
            return False
        token = self.tokens[self.index + 1]
        return token.kind == kind and (value is None or token.value == value)

    def advance(self) -> Token:
        token = self.peek()
        self.index += 1
        return token

    def previous(self) -> Token:
        return self.tokens[self.index - 1]

    def peek(self) -> Token:
        return self.tokens[self.index]

    def at_end(self) -> bool:
        return self.peek().kind == "EOF"


SUPPORTED_FUNCTIONS = {"SUM", "AVERAGE", "MIN", "MAX", "COUNT"}


TOKEN_RE = re.compile(
    r"""
    (?P<SPACE>\s+)
  | (?P<NUMBER>(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?)
  | (?P<STRING>"(?:[^"]|"")*")
  | (?P<SHEET>'(?:[^']|'')+')
  | (?P<WORKBOOK>\[[^\]]+\])
  | (?P<CELL>\$?[A-Za-z]{1,3}\$?\d+)
  | (?P<COLUMN>\$?[A-Za-z]{1,3}(?=\s*[:),]))
  | (?P<IDENT>[_A-Za-z][._A-Za-z0-9]*)
  | (?P<OP><>|<=|>=|[=+\-*/^(),:!%<>])
    """,
    re.VERBOSE,
)


def tokenize(formula: str) -> list[Token]:
    tokens: list[Token] = []
    position = 0
    while position < len(formula):
        match = TOKEN_RE.match(formula, position)
        if not match:
            raise FormulaParseError(f"Unsupported character {formula[position]!r} at {position}")
        kind = match.lastgroup or ""
        raw_value = match.group(kind)
        if kind != "SPACE":
            value = raw_value
            if kind == "STRING":
                value = raw_value[1:-1].replace('""', '"')
            elif kind == "SHEET":
                value = raw_value[1:-1].replace("''", "'")
            tokens.append(Token(kind, value, position))
        position = match.end()
    tokens.append(Token("EOF", "", position))
    return tokens


def evaluate_function(function_name: str, args: list[Any], formula: str) -> Any:
    values = list(flatten(args))
    sentinel = first_sentinel(values)
    if sentinel is not None:
        return sentinel

    numeric_values = [to_number(value, formula) for value in values if can_be_number(value)]
    if function_name == "SUM":
        return sum(numeric_values)
    if function_name == "AVERAGE":
        if not numeric_values:
            return FormulaSentinel("error", "division_by_zero", formula=formula)
        return sum(numeric_values) / len(numeric_values)
    if function_name == "MIN":
        return min(numeric_values) if numeric_values else 0
    if function_name == "MAX":
        return max(numeric_values) if numeric_values else 0
    if function_name == "COUNT":
        return len(numeric_values)
    return FormulaSentinel("unsupported", "unsupported_function", formula=formula, detail=function_name)


def apply_binary(operator: str, left: Any, right: Any, formula: str) -> Any:
    sentinel = first_sentinel([left, right])
    if sentinel is not None:
        return sentinel
    left_number = to_number(left, formula)
    right_number = to_number(right, formula)
    if operator == "+":
        return left_number + right_number
    if operator == "-":
        return left_number - right_number
    if operator == "*":
        return left_number * right_number
    if operator == "/":
        if right_number == 0:
            raise ZeroDivisionError
        return left_number / right_number
    if operator == "^":
        return left_number**right_number
    raise FormulaParseError(f"Unsupported operator {operator!r}")


def flatten(values: list[Any]) -> list[Any]:
    flattened: list[Any] = []
    for value in values:
        if isinstance(value, list):
            flattened.extend(flatten(value))
        else:
            flattened.append(value)
    return flattened


def first_sentinel(values: list[Any]) -> FormulaSentinel | None:
    for value in values:
        if isinstance(value, FormulaSentinel):
            return value
    return None


def can_be_number(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, bool | FormulaSentinel):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text == "":
            return True
        try:
            float(text)
        except ValueError:
            return False
        return True
    return False


def to_number(value: Any, formula: str) -> float:
    if isinstance(value, FormulaSentinel):
        raise FormulaParseError(value.reason)
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return float(value.toordinal())
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text == "":
            return 0.0
        try:
            return float(text)
        except ValueError as exc:
            raise FormulaParseError(f"Cannot coerce {value!r} to number in {formula!r}") from exc
    raise FormulaParseError(f"Cannot coerce {type(value).__name__} to number in {formula!r}")


def parse_number(value: str) -> int | float:
    parsed = float(value)
    return int(parsed) if parsed.is_integer() and "." not in value and "e" not in value.lower() else parsed


def normalize_coordinate(coordinate: str) -> str:
    return strip_dollars(coordinate).upper()


def normalize_sheet_reference_name(sheet_name: str) -> str:
    return re.sub(r"^\[[^\]]+\]", "", sheet_name)


def strip_dollars(reference: str) -> str:
    return reference.replace("$", "")


def is_formula_sentinel(value: Any) -> bool:
    return isinstance(value, FormulaSentinel)
