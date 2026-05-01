from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
import json
from pathlib import Path
from typing import Any, Literal

import polars as pl
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from pydantic import BaseModel, Field, model_validator


RAW_DATA_SHEETS = [
    "RAW DATA_Master File",
    "RAW DATA_COGS & Freight",
    "RAW DATA_Payroll",
]


class CellWrite(BaseModel):
    """Configurable target-cell write for the consolidated P&L template."""

    sheet_name: str
    cell: str
    source_key: str | None = None
    value: Any = None
    formula: str | None = None
    formula_template: str | None = None
    formula_source_keys: dict[str, str] = Field(default_factory=dict)
    formula_missing_value: Any = None
    clear: bool = False
    required: bool = False
    value_type: Literal["auto", "number", "string", "formula", "blank"] = "auto"
    overwrite_formula: bool = False
    note: str = ""

    @model_validator(mode="after")
    def validate_source(self) -> "CellWrite":
        sources = [
            self.source_key is not None,
            self.value is not None,
            self.formula is not None,
            self.formula_template is not None,
            self.clear,
        ]
        if sum(sources) != 1:
            raise ValueError(
                "Exactly one of source_key, value, formula_template, formula, or clear must be configured"
            )
        if self.formula_template is None and self.formula_source_keys:
            raise ValueError("formula_source_keys requires formula_template")
        return self


class SheetVisibility(BaseModel):
    sheet_name: str
    state: Literal["visible", "hidden", "veryHidden"]


class CellValidation(BaseModel):
    """Post-write validation against the workbook cell value before recalculation."""

    name: str
    sheet_name: str
    cell: str
    expected_source_key: str | None = None
    expected_value: Any = None
    tolerance: float = 0.01

    @model_validator(mode="after")
    def validate_expected(self) -> "CellValidation":
        if self.expected_source_key is None and self.expected_value is None:
            raise ValueError("expected_source_key or expected_value is required")
        return self


class ConsolidatedPNLWriterConfig(BaseModel):
    """Agent-adjustable writer config for the HGF consolidated P&L template."""

    raw_data_sheets: list[str] = Field(default_factory=lambda: list(RAW_DATA_SHEETS))
    cell_writes: list[CellWrite] = Field(default_factory=list)
    validations: list[CellValidation] = Field(default_factory=list)
    sheet_visibility: list[SheetVisibility] = Field(default_factory=list)
    preserve_existing_formulas: bool = True
    allow_formula_values: bool = True
    fail_on_missing_required_values: bool = True
    set_full_calc_on_load: bool = True
    calc_mode: Literal["auto", "manual", "autoNoTable"] = "auto"

    @classmethod
    def from_json_file(cls, path: Path | None) -> "ConsolidatedPNLWriterConfig":
        if path is None:
            return default_consolidated_pnl_writer_config()
        return cls.model_validate_json(path.read_text(encoding="utf-8"))


@dataclass
class ConsolidatedPNLWriteResult:
    template_path: Path
    output_path: Path
    written_cells: list[dict[str, Any]]
    skipped_cells: list[dict[str, Any]] = field(default_factory=list)
    validation_results: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def writes(self) -> pl.DataFrame:
        return pl.DataFrame(self.written_cells, infer_schema_length=None)

    @property
    def skips(self) -> pl.DataFrame:
        return pl.DataFrame(self.skipped_cells, infer_schema_length=None)

    @property
    def validations(self) -> pl.DataFrame:
        return pl.DataFrame(self.validation_results, infer_schema_length=None)

    def to_dict(self) -> dict[str, Any]:
        return {
            "template_path": str(self.template_path),
            "output_path": str(self.output_path),
            "warnings": self.warnings,
            "written_cells": self.written_cells,
            "skipped_cells": self.skipped_cells,
            "validation_results": self.validation_results,
        }


def write_consolidated_pnl(
    template_path: Path,
    output_path: Path,
    values: dict[str, Any] | None = None,
    config: ConsolidatedPNLWriterConfig | None = None,
) -> ConsolidatedPNLWriteResult:
    config = config or default_consolidated_pnl_writer_config()
    values = values or {}
    workbook = load_workbook(template_path)
    written_cells: list[dict[str, Any]] = []
    skipped_cells: list[dict[str, Any]] = []
    warnings: list[str] = []

    try:
        validate_sheets_exist(workbook, config.raw_data_sheets, warnings)
        apply_sheet_visibility(workbook, config.sheet_visibility, warnings)

        for write in config.cell_writes:
            applied = apply_cell_write(workbook, write, values, config, warnings)
            if applied["status"] == "written":
                written_cells.append(applied)
            else:
                skipped_cells.append(applied)

        validation_results = run_validations(workbook, config.validations, values, warnings)
        if config.set_full_calc_on_load:
            configure_recalculation(workbook, config)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()

    return ConsolidatedPNLWriteResult(
        template_path=template_path,
        output_path=output_path,
        written_cells=written_cells,
        skipped_cells=skipped_cells,
        validation_results=validation_results,
        warnings=warnings,
    )


def apply_cell_write(
    workbook: Workbook,
    write: CellWrite,
    values: dict[str, Any],
    config: ConsolidatedPNLWriterConfig,
    warnings: list[str],
) -> dict[str, Any]:
    if write.sheet_name not in workbook.sheetnames:
        warning = f"Sheet not found for write: {write.sheet_name}"
        warnings.append(warning)
        return write_result(write, "skipped", warning, None, None)

    ws = workbook[write.sheet_name]
    cell = ws[write.cell]
    old_value = cell.value
    if (
        config.preserve_existing_formulas
        and is_formula(old_value)
        and not write.overwrite_formula
    ):
        warning = f"Skipped formula cell {write.sheet_name}!{write.cell}"
        warnings.append(warning)
        return write_result(write, "skipped", warning, old_value, None)

    try:
        new_value = resolve_write_value(write, values, config)
    except KeyError as exc:
        warning = str(exc)
        if config.fail_on_missing_required_values and write.required:
            warnings.append(warning)
            raise
        if write.required:
            warnings.append(warning)
        return write_result(write, "skipped", warning, old_value, None)

    cell.value = new_value
    return write_result(write, "written", None, old_value, new_value)


def resolve_write_value(
    write: CellWrite,
    values: dict[str, Any],
    config: ConsolidatedPNLWriterConfig,
) -> Any:
    if write.clear:
        return None
    if write.formula is not None:
        return normalize_formula(write.formula)
    if write.formula_template is not None:
        return resolve_formula_template(write, values, config)
    if write.value is not None:
        return coerce_value(write.value, write.value_type, config)
    if write.source_key is None:
        raise KeyError(f"No source key configured for {write.sheet_name}!{write.cell}")

    resolved = resolve_source_key(values, write.source_key)
    if resolved is None and write.required:
        raise KeyError(f"Missing required value: {write.source_key}")
    return coerce_value(resolved, write.value_type, config)


def resolve_formula_template(
    write: CellWrite,
    values: dict[str, Any],
    config: ConsolidatedPNLWriterConfig,
) -> str:
    replacements: dict[str, Any] = {}
    resolved_sources = 0
    missing_source_keys: list[str] = []
    for placeholder, source_key in write.formula_source_keys.items():
        try:
            value = resolve_source_key(values, source_key)
            resolved_sources += 1
        except KeyError:
            missing_source_keys.append(source_key)
            if write.formula_missing_value is None:
                if write.required:
                    raise
                value = 0
            else:
                value = write.formula_missing_value
        replacements[placeholder] = coerce_value(value, "number", config)

    if write.formula_source_keys and resolved_sources == 0:
        raise KeyError(
            f"Missing all formula source values for {write.sheet_name}!{write.cell}: "
            f"{', '.join(missing_source_keys)}"
        )

    try:
        formula = write.formula_template.format(**replacements)
    except KeyError as exc:
        raise KeyError(f"Missing formula placeholder for {write.sheet_name}!{write.cell}: {exc}") from exc
    return normalize_formula(formula)


def resolve_source_key(values: dict[str, Any], source_key: str) -> Any:
    if source_key in values:
        return values[source_key]

    current: Any = values
    for part in source_key.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            raise KeyError(f"Missing value: {source_key}")
    return current


def coerce_value(
    value: Any,
    value_type: Literal["auto", "number", "string", "formula", "blank"],
    config: ConsolidatedPNLWriterConfig,
) -> Any:
    if value_type == "blank":
        return None
    if value is None:
        return None
    if isinstance(value, Decimal):
        value = float(value)
    if value_type == "formula":
        return normalize_formula(value)
    if isinstance(value, str) and value.startswith("="):
        if not config.allow_formula_values:
            raise ValueError(f"Formula value is not allowed: {value}")
        return value
    if value_type == "number":
        return to_float(value)
    if value_type == "string":
        return str(value)
    return value


def normalize_formula(value: Any) -> str:
    formula = str(value)
    return formula if formula.startswith("=") else f"={formula}"


def write_result(
    write: CellWrite,
    status: str,
    warning: str | None,
    old_value: Any,
    new_value: Any,
) -> dict[str, Any]:
    return {
        "sheet_name": write.sheet_name,
        "cell": write.cell,
        "source_key": write.source_key,
        "status": status,
        "old_value": serializable_value(old_value),
        "new_value": serializable_value(new_value),
        "warning": warning,
        "note": write.note,
    }


def run_validations(
    workbook: Workbook,
    validations: list[CellValidation],
    values: dict[str, Any],
    warnings: list[str],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for validation in validations:
        if validation.sheet_name not in workbook.sheetnames:
            warning = f"Sheet not found for validation: {validation.sheet_name}"
            warnings.append(warning)
            results.append(validation_result(validation, "missing_sheet", warning, None, None, None))
            continue

        actual = workbook[validation.sheet_name][validation.cell].value
        try:
            expected = (
                resolve_source_key(values, validation.expected_source_key)
                if validation.expected_source_key is not None
                else validation.expected_value
            )
        except KeyError as exc:
            warning = str(exc)
            warnings.append(warning)
            results.append(validation_result(validation, "missing_expected", warning, actual, None, None))
            continue

        difference = numeric_difference(actual, expected)
        if difference is None:
            ok = actual == expected
        else:
            ok = abs(difference) <= validation.tolerance
        status = "ok" if ok else "mismatch"
        warning = None
        if not ok:
            warning = f"Validation failed: {validation.name}"
            warnings.append(warning)
        results.append(validation_result(validation, status, warning, actual, expected, difference))
    return results


def validation_result(
    validation: CellValidation,
    status: str,
    warning: str | None,
    actual: Any,
    expected: Any,
    difference: float | None,
) -> dict[str, Any]:
    return {
        "name": validation.name,
        "sheet_name": validation.sheet_name,
        "cell": validation.cell,
        "expected_source_key": validation.expected_source_key,
        "status": status,
        "actual": serializable_value(actual),
        "expected": serializable_value(expected),
        "difference": difference,
        "tolerance": validation.tolerance,
        "warning": warning,
    }


def apply_sheet_visibility(
    workbook: Workbook,
    visibility_rules: list[SheetVisibility],
    warnings: list[str],
) -> None:
    for rule in visibility_rules:
        if rule.sheet_name not in workbook.sheetnames:
            warnings.append(f"Sheet not found for visibility rule: {rule.sheet_name}")
            continue
        workbook[rule.sheet_name].sheet_state = rule.state


def validate_sheets_exist(workbook: Workbook, sheet_names: list[str], warnings: list[str]) -> None:
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            warnings.append(f"Expected raw data sheet is missing: {sheet_name}")


def configure_recalculation(workbook: Workbook, config: ConsolidatedPNLWriterConfig) -> None:
    workbook.calculation.calcMode = config.calc_mode
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True


def default_consolidated_pnl_writer_config() -> ConsolidatedPNLWriterConfig:
    return ConsolidatedPNLWriterConfig(
        cell_writes=default_raw_data_cell_writes(),
        validations=default_raw_data_validations(),
        sheet_visibility=[
            SheetVisibility(sheet_name="MARCH 2026 FULL ", state="visible"),
            SheetVisibility(sheet_name="RAW DATA_Master File", state="hidden"),
            SheetVisibility(sheet_name="RAW DATA_COGS & Freight", state="hidden"),
            SheetVisibility(sheet_name="RAW DATA_Payroll", state="hidden"),
        ],
    )


def default_raw_data_cell_writes() -> list[CellWrite]:
    """Known March template staging-cell targets.

    Source keys are intentionally semantic rather than tied to one extractor. The agent
    can build this value map after validating extractor outputs and overrides.
    """

    return [
        # Hidden source-total cells on the full report without raw-tab staging.
        map_cell(
            "MARCH 2026 FULL ",
            "EB48",
            "full_report.source_totals.employee_benefits",
        ),
        # Visible Payroll Art/IT actual formulas. These rows are manually refreshed
        # from the Payroll sheet's department allocation totals each month.
        formula_cell(
            "MARCH 2026 FULL ",
            "E44",
            "{direct}+F7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.trend_house",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art Trend House: direct TH allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "P44",
            "{direct}+Q7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.og_specialty_usa",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art OG Specialty: direct B&M USA allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AA44",
            "{direct}+AB7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.online_lux",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art Online Lux: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AL44",
            "{direct}+AM7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.online",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art Online: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AW44",
            "{direct}+AX7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.dtc",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art OG-DTC: direct OG DTC allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "BH44",
            "{direct}+BI7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.all_pop_art",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art All Pop Art: direct APA allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "BV44",
            "{direct}+BW7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.art.ink",
                "general": "raw_payroll.allocation_breakdowns.art.general",
            },
            note="Payroll Art Ink: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "E45",
            "{direct}+F7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.trend_house",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT Trend House: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "P45",
            "{direct}+Q7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.og_specialty_usa",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT OG Specialty: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AA45",
            "{direct}+AB7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.online_lux",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT Online Lux: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AL45",
            "{direct}+AM7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.online",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT Online: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "AW45",
            "{direct}+AX7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.dtc",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT OG-DTC: direct OG DTC allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "BH45",
            "{direct}+BI7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.all_pop_art",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT All Pop Art: direct allocation plus revenue-share General allocation.",
        ),
        formula_cell(
            "MARCH 2026 FULL ",
            "BV45",
            "{direct}+BW7*{general}",
            {
                "direct": "raw_payroll.allocation_breakdowns.it.ink",
                "general": "raw_payroll.allocation_breakdowns.it.general",
            },
            note="Payroll IT Ink: direct allocation plus revenue-share General allocation.",
        ),
        # RAW DATA_Master File: GL/account totals and allocation blocks.
        map_cell("RAW DATA_Master File", "B6", "raw_master.gl.warehouse_rent_curci"),
        map_cell("RAW DATA_Master File", "B7", "raw_master.gl.payroll_processing"),
        map_cell("RAW DATA_Master File", "B8", "raw_master.gl.consulting_expense"),
        map_cell("RAW DATA_Master File", "B9", "raw_master.gl.rent_office"),
        map_cell("RAW DATA_Master File", "B10", "raw_master.gl.rent_showrooms"),
        map_cell("RAW DATA_Master File", "B11", "raw_master.gl.equipment_lease"),
        map_cell("RAW DATA_Master File", "D11", "raw_master.gl.equipment_lease_adjustment"),
        map_cell("RAW DATA_Master File", "B12", "raw_master.gl.repairs_maintenance"),
        map_cell("RAW DATA_Master File", "B14", "raw_master.gl.vehicle_expense"),
        map_cell("RAW DATA_Master File", "B16", "raw_master.gl.advertising_marketing"),
        map_cell("RAW DATA_Master File", "B17", "raw_master.gl.office_supplies"),
        map_cell("RAW DATA_Master File", "B18", "raw_master.gl.art_assets"),
        map_cell("RAW DATA_Master File", "B19", "raw_master.gl.employee_nurturing"),
        map_cell("RAW DATA_Master File", "B21", "raw_master.gl.software_web_services"),
        map_cell("RAW DATA_Master File", "B22", "raw_master.gl.bank_fees"),
        map_cell("RAW DATA_Master File", "D22", "raw_master.gl.bank_fees_adjustment"),
        map_cell("RAW DATA_Master File", "B23", "raw_master.gl.merchant_account_fees"),
        map_cell("RAW DATA_Master File", "D23", "raw_master.gl.merchant_account_fees_adjustment"),
        map_cell("RAW DATA_Master File", "B24", "raw_master.gl.cleaning_janitorial"),
        map_cell("RAW DATA_Master File", "B25", "raw_master.gl.hr_recruiting"),
        map_cell("RAW DATA_Master File", "B26", "raw_master.gl.hr_training"),
        map_cell("RAW DATA_Master File", "B27", "raw_master.gl.insurance"),
        map_cell("RAW DATA_Master File", "B28", "raw_master.gl.telephone_internet"),
        map_cell("RAW DATA_Master File", "B29", "raw_master.gl.utilities"),
        map_cell("RAW DATA_Master File", "B30", "raw_master.gl.licenses_taxes_permits"),
        map_cell("RAW DATA_Master File", "B32", "raw_master.gl.dues_subscriptions"),
        map_cell("RAW DATA_Master File", "B33", "raw_master.gl.travel"),
        map_cell("RAW DATA_Master File", "B34", "raw_master.gl.meals_entertainment"),
        map_cell("RAW DATA_Master File", "B36", "raw_master.gl.professional_fees_accounting"),
        map_cell("RAW DATA_Master File", "B37", "raw_master.gl.professional_fees_legal"),
        map_cell("RAW DATA_Master File", "B38", "raw_master.gl.professional_fees_it"),
        map_cell("RAW DATA_Master File", "B41", "raw_master.gl.loc_interest"),
        map_cell("RAW DATA_Master File", "B46", "raw_master.software_web.corp"),
        map_cell("RAW DATA_Master File", "B47", "raw_master.software_web.online"),
        map_cell("RAW DATA_Master File", "B48", "raw_master.software_web.dtc"),
        map_cell("RAW DATA_Master File", "B49", "raw_master.software_web.online_lux"),
        map_cell("RAW DATA_Master File", "B52", "raw_master.consulting.corp"),
        map_cell("RAW DATA_Master File", "B53", "raw_master.consulting.trend_house"),
        map_cell("RAW DATA_Master File", "B54", "raw_master.consulting.dtc"),
        map_cell("RAW DATA_Master File", "B55", "raw_master.consulting.online"),
        map_cell("RAW DATA_Master File", "B56", "raw_master.consulting.online_lux"),
        map_cell("RAW DATA_Master File", "B59", "raw_master.advertising.dtc"),
        map_cell("RAW DATA_Master File", "B60", "raw_master.advertising.online"),
        map_cell("RAW DATA_Master File", "B61", "raw_master.advertising.online_lux"),
        map_cell("RAW DATA_Master File", "B62", "raw_master.advertising.trend_house"),
        map_cell("RAW DATA_Master File", "B71", "raw_master.sales.online"),
        map_cell("RAW DATA_Master File", "B72", "raw_master.sales.dtc"),
        map_cell("RAW DATA_Master File", "B73", "raw_master.sales.apa"),
        map_cell("RAW DATA_Master File", "B74", "raw_master.sales.ink"),
        map_cell("RAW DATA_Master File", "B77", "raw_master.returns.trend_house"),
        map_cell("RAW DATA_Master File", "B78", "raw_master.returns.og_specialty_usa"),
        map_cell("RAW DATA_Master File", "B79", "raw_master.returns.og_specialty_trade"),
        map_cell("RAW DATA_Master File", "B81", "raw_master.returns.dtc"),
        map_cell("RAW DATA_Master File", "B82", "raw_master.returns.apa"),
        map_cell("RAW DATA_Master File", "B83", "raw_master.returns.ink"),
        map_cell("RAW DATA_Master File", "B120", "raw_master.travel.corp"),
        map_cell("RAW DATA_Master File", "B121", "raw_master.travel.dtc"),
        map_cell("RAW DATA_Master File", "B122", "raw_master.travel.trend_house"),
        map_cell("RAW DATA_Master File", "B126", "raw_master.meals.corp"),
        map_cell("RAW DATA_Master File", "B127", "raw_master.meals.dtc"),
        map_cell("RAW DATA_Master File", "B128", "raw_master.meals.trend_house"),
        map_cell("RAW DATA_Master File", "B131", "raw_master.samples.specialty_trade"),
        map_cell("RAW DATA_Master File", "B133", "raw_master.samples.dtc"),
        # RAW DATA_COGS & Freight.
        map_cell("RAW DATA_COGS & Freight", "E2", "raw_cogs.current_month.cogs.bm_usa_samples"),
        map_cell("RAW DATA_COGS & Freight", "G2", "raw_cogs.current_month.cogs.online"),
        map_cell("RAW DATA_COGS & Freight", "K2", "raw_cogs.current_month.cogs.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "L2", "raw_cogs.current_month.cogs.online_textiles_mww"),
        map_cell("RAW DATA_COGS & Freight", "P2", "raw_cogs.current_month.cogs.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q2", "raw_cogs.current_month.cogs.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R2", "raw_cogs.current_month.cogs.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "E3", "raw_cogs.material.bm_usa_samples"),
        map_cell("RAW DATA_COGS & Freight", "K3", "raw_cogs.material.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "P3", "raw_cogs.material.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q3", "raw_cogs.material.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R3", "raw_cogs.material.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "E4", "raw_cogs.labor.bm_usa_samples"),
        map_cell("RAW DATA_COGS & Freight", "K4", "raw_cogs.labor.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "P4", "raw_cogs.labor.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q4", "raw_cogs.labor.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R4", "raw_cogs.labor.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "G5", "raw_cogs.shipping_actual.online"),
        map_cell("RAW DATA_COGS & Freight", "K5", "raw_cogs.shipping_actual.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "P5", "raw_cogs.shipping_actual.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q5", "raw_cogs.shipping_actual.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R5", "raw_cogs.shipping_actual.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "G6", "raw_cogs.fedex.online"),
        map_cell("RAW DATA_COGS & Freight", "K6", "raw_cogs.fedex.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "P6", "raw_cogs.fedex.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q6", "raw_cogs.fedex.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R6", "raw_cogs.fedex.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "T6", "raw_cogs.fedex.corporate_shipping"),
        map_cell("RAW DATA_COGS & Freight", "G7", "raw_cogs.ups.online"),
        map_cell("RAW DATA_COGS & Freight", "K7", "raw_cogs.ups.online_usa"),
        map_cell("RAW DATA_COGS & Freight", "P7", "raw_cogs.ups.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "Q7", "raw_cogs.ups.og_dtc_returns"),
        map_cell("RAW DATA_COGS & Freight", "R7", "raw_cogs.ups.all_pop_art"),
        map_cell("RAW DATA_COGS & Freight", "P8", "raw_cogs.shipping_quoted_dtc.og_dtc"),
        map_cell("RAW DATA_COGS & Freight", "B12", "raw_cogs.shipping_for_samples.current_month"),
        map_cell("RAW DATA_COGS & Freight", "C27", "raw_cogs.trend_house.total.revenue"),
        map_cell("RAW DATA_COGS & Freight", "D27", "raw_cogs.trend_house.total.production_cost"),
        map_cell("RAW DATA_COGS & Freight", "E27", "raw_cogs.trend_house.total.shipping_cost"),
        map_cell("RAW DATA_COGS & Freight", "E29", "raw_cogs.tariffs"),
        # RAW DATA_Payroll.
        map_cell("RAW DATA_Payroll", "B2", "raw_payroll.sales.trend_house"),
        map_cell("RAW DATA_Payroll", "B3", "raw_payroll.sales.og_specialty_usa"),
        map_cell("RAW DATA_Payroll", "B4", "raw_payroll.sales.online_lux"),
        map_cell("RAW DATA_Payroll", "B5", "raw_payroll.sales.online"),
        map_cell("RAW DATA_Payroll", "B6", "raw_payroll.sales.dtc"),
        map_cell("RAW DATA_Payroll", "B9", "raw_payroll.corp.corp"),
        map_cell("RAW DATA_Payroll", "B10", "raw_payroll.corp.art"),
        map_cell("RAW DATA_Payroll", "B11", "raw_payroll.corp.it"),
        map_cell("RAW DATA_Payroll", "B14", "raw_payroll.production"),
        map_cell("RAW DATA_Payroll", "B21", "raw_payroll.lital_allocation.dtc"),
        map_cell("RAW DATA_Payroll", "B22", "raw_payroll.lital_allocation.online"),
        map_cell("RAW DATA_Payroll", "B23", "raw_payroll.lital_allocation.trend_house"),
        CellWrite(
            sheet_name="RAW DATA_Payroll",
            cell="A24",
            value="CORP",
            value_type="string",
            note="March 2026 final workbook moves CORP to row 24 in the Lital allocation block.",
        ),
        map_cell("RAW DATA_Payroll", "B24", "raw_payroll.lital_allocation.corp"),
        CellWrite(
            sheet_name="RAW DATA_Payroll",
            cell="A25",
            clear=True,
            note="Clear unused Lital allocation label row after moving CORP to row 24.",
        ),
        CellWrite(
            sheet_name="RAW DATA_Payroll",
            cell="B25",
            formula="SUM(B21:B24)",
            note="Total the active Lital allocation rows.",
        ),
        CellWrite(
            sheet_name="RAW DATA_Payroll",
            cell="B26",
            clear=True,
            overwrite_formula=True,
            note="Clear the template's old Lital allocation total row.",
        ),
    ]


def default_raw_data_validations() -> list[CellValidation]:
    return [
        CellValidation(
            name="raw payroll production",
            sheet_name="RAW DATA_Payroll",
            cell="B14",
            expected_source_key="raw_payroll.production",
        ),
        CellValidation(
            name="raw master DTC sales",
            sheet_name="RAW DATA_Master File",
            cell="B72",
            expected_source_key="raw_master.sales.dtc",
        ),
        CellValidation(
            name="raw COGS Trend House revenue",
            sheet_name="RAW DATA_COGS & Freight",
            cell="C27",
            expected_source_key="raw_cogs.trend_house.total.revenue",
        ),
    ]


def map_cell(sheet_name: str, cell: str, source_key: str) -> CellWrite:
    return CellWrite(
        sheet_name=sheet_name,
        cell=cell,
        source_key=source_key,
        value_type="number",
        required=False,
    )


def formula_cell(
    sheet_name: str,
    cell: str,
    formula_template: str,
    source_keys: dict[str, str],
    note: str,
) -> CellWrite:
    return CellWrite(
        sheet_name=sheet_name,
        cell=cell,
        formula_template=formula_template,
        formula_source_keys=source_keys,
        formula_missing_value=0,
        value_type="formula",
        overwrite_formula=True,
        note=note,
    )


def load_values_json(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def write_default_config(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(default_consolidated_pnl_writer_config().model_dump(), indent=2) + "\n",
        encoding="utf-8",
    )


def write_example_values_from_workbook(
    completed_workbook_path: Path,
    output_path: Path,
    config: ConsolidatedPNLWriterConfig | None = None,
) -> dict[str, Any]:
    """Write a values JSON by reading configured source cells from a completed workbook."""

    config = config or default_consolidated_pnl_writer_config()
    workbook = load_workbook(completed_workbook_path, data_only=False)
    values: dict[str, Any] = {}
    try:
        for write in config.cell_writes:
            if write.source_key is None or write.sheet_name not in workbook.sheetnames:
                continue
            value = workbook[write.sheet_name][write.cell].value
            if value in (None, ""):
                continue
            set_nested_value(values, write.source_key, serializable_value(value))
    finally:
        workbook.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(values, indent=2) + "\n", encoding="utf-8")
    return values


def set_nested_value(values: dict[str, Any], source_key: str, value: Any) -> None:
    current = values
    parts = source_key.split(".")
    for part in parts[:-1]:
        next_value = current.setdefault(part, {})
        if not isinstance(next_value, dict):
            raise ValueError(f"Cannot set nested source key through scalar: {source_key}")
        current = next_value
    current[parts[-1]] = value


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float | Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("$", "")
        if text in {"", "-", "N/A", "#N/A"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def numeric_difference(actual: Any, expected: Any) -> float | None:
    actual_float = to_float(actual)
    expected_float = to_float(expected)
    if actual_float is None or expected_float is None:
        return None
    return actual_float - expected_float


def serializable_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value
