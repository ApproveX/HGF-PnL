from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

import typer
from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table


app = typer.Typer(help="Explore HGF close-package Excel workbooks.")
console = Console()


CATEGORY_TERMS: dict[str, tuple[str, ...]] = {
    "payroll": ("payroll", "earning", "salary", "bonus", "wages", "hourly"),
    "consolidated_pnl": ("consolidated", "full company", "profit and loss", "p&l"),
    "department_pnl": ("departmental", "department", "dept"),
    "budget": ("budget", "forecast", "projection", "proforma"),
    "cogs_inventory": ("cogs", "inventory", "production inventory"),
    "revenue_sales": ("revenue", "sales", "dtc", "shopify", "trend house"),
    "tax": ("tax",),
    "royalty": ("royalty",),
    "chargeback": ("chargeback", "deduction"),
    "ar_ap_balance": ("aging", "accts rec", "accts pay", "balance sheet"),
    "kpi": ("kpi",),
}


def classify_workbook(path: Path, sheet_names: list[str]) -> str:
    haystack = " ".join([str(path).lower(), *(name.lower() for name in sheet_names)])
    scores = {
        category: sum(term in haystack for term in terms)
        for category, terms in CATEGORY_TERMS.items()
    }
    category, score = max(scores.items(), key=lambda item: item[1])
    return category if score else "unclassified"


def sample_rows(ws: Any, max_rows: int = 3, scan_rows: int = 20, scan_cols: int = 12) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 0, scan_rows),
            max_col=min(ws.max_column or 0, scan_cols),
            values_only=True,
        ),
        start=1,
    ):
        values = ["" if value is None else str(value).replace("\n", " ")[:120] for value in row]
        if any(values):
            rows.append({"row": row_idx, "values": values})
        if len(rows) >= max_rows:
            break
    return rows


def inspect_workbook(path: Path, include_samples: bool) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet_infos = []
        total_formula_count = 0
        for ws in workbook.worksheets:
            formula_count = 0
            fill_counts: Counter[str] = Counter()
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formula_count += 1
                    if cell.value not in (None, ""):
                        fill = cell.fill.fgColor
                        color = fill.rgb if fill.type == "rgb" else None
                        if color and color not in {"00000000", "FFFFFFFF"}:
                            fill_counts[color] += 1

            total_formula_count += formula_count
            info: dict[str, Any] = {
                "name": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "formula_count": formula_count,
                "fill_colors": dict(fill_counts.most_common(12)),
            }
            if include_samples:
                info["sample_rows"] = sample_rows(ws)
            sheet_infos.append(info)

        sheet_names = [sheet["name"] for sheet in sheet_infos]
        return {
            "path": str(path),
            "size_kb": round(path.stat().st_size / 1024, 1),
            "category": classify_workbook(path, sheet_names),
            "sheet_count": len(sheet_infos),
            "formula_count": total_formula_count,
            "sheets": sheet_infos,
        }
    finally:
        workbook.close()


@app.command()
def scan(
    root: Path = typer.Argument(Path("sample_files"), help="Folder containing close-package files."),
    json_out: Path | None = typer.Option(None, "--json-out", help="Optional JSON output path."),
    samples: bool = typer.Option(True, "--samples/--no-samples", help="Include sample rows."),
) -> None:
    workbooks = sorted(path for path in root.rglob("*.xlsx") if ":Zone.Identifier" not in path.name)
    results = [inspect_workbook(path, include_samples=samples) for path in workbooks]

    if json_out:
        json_out.parent.mkdir(parents=True, exist_ok=True)
        json_out.write_text(json.dumps(results, indent=2), encoding="utf-8")
        console.print(f"Wrote {json_out}")

    category_counts = Counter(result["category"] for result in results)
    console.print(f"Scanned {len(results)} workbooks under {root}")
    console.print("Categories: " + ", ".join(f"{key}={value}" for key, value in category_counts.items()))

    table = Table(title="Largest Workbooks")
    table.add_column("KB", justify="right")
    table.add_column("Sheets", justify="right")
    table.add_column("Category")
    table.add_column("Path")
    for result in sorted(results, key=lambda item: item["size_kb"], reverse=True)[:20]:
        table.add_row(
            f'{result["size_kb"]:,.1f}',
            str(result["sheet_count"]),
            result["category"],
            result["path"],
        )
    console.print(table)


if __name__ == "__main__":
    app()
