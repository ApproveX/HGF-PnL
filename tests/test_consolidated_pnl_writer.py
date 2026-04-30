from pathlib import Path

from openpyxl import load_workbook

from hgf_pnl.writers.consolidated_pnl import (
    CellValidation,
    CellWrite,
    ConsolidatedPNLWriterConfig,
    default_consolidated_pnl_writer_config,
    write_consolidated_pnl,
)


TEMPLATE = Path("sample_files/HGF CONSOLIDATED_MARCH 2026 Template.xlsx")
FINAL = Path("sample_files/P&L_S/FULL COMPANY P&L_s/HGF CONSOLIDATED_MARCH 2026_FINAL.xlsx")


def test_writes_selected_raw_data_values_against_march_example(tmp_path: Path) -> None:
    output = tmp_path / "consolidated.xlsx"
    values = {
        "raw_master": {
            "gl": {"consulting_expense": 118_604.51},
            "sales": {"dtc": 163_408.05},
            "travel": {"dtc": 4_172.14},
            "meals": {"trend_house": 3_006.68},
        },
        "raw_cogs": {
            "trend_house": {"total": {"revenue": 1_242_393.40}},
            "tariffs": 73_451.30,
        },
        "raw_payroll": {
            "sales": {"trend_house": 36_909.05},
            "production": 64_635.26,
            "lital_allocation": {
                "dtc": 772,
                "online": 6_176,
                "trend_house": 1_544,
                "corp": 6_948,
            },
        },
        "full_report": {
            "source_totals": {
                "employee_benefits": 10_897.54,
            },
        },
    }

    result = write_consolidated_pnl(TEMPLATE, output, values, default_consolidated_pnl_writer_config())

    assert output.exists()
    assert not result.warnings
    assert {row["status"] for row in result.validation_results} == {"ok"}

    generated = load_workbook(output, data_only=False)
    final = load_workbook(FINAL, data_only=False)
    try:
        for sheet, cell in [
            ("RAW DATA_Master File", "B8"),
            ("RAW DATA_Master File", "B72"),
            ("RAW DATA_Master File", "B121"),
            ("RAW DATA_Master File", "B128"),
            ("RAW DATA_COGS & Freight", "C27"),
            ("RAW DATA_COGS & Freight", "E29"),
            ("RAW DATA_Payroll", "B2"),
            ("RAW DATA_Payroll", "B14"),
            ("RAW DATA_Payroll", "A24"),
            ("RAW DATA_Payroll", "B24"),
            ("RAW DATA_Payroll", "A25"),
            ("RAW DATA_Payroll", "B25"),
            ("RAW DATA_Payroll", "B26"),
            ("MARCH 2026 FULL ", "EB48"),
        ]:
            assert generated[sheet][cell].value == final[sheet][cell].value

        assert generated["MARCH 2026 FULL "]["E8"].value == "=+'RAW DATA_Master File'!B68"
        assert generated.calculation.fullCalcOnLoad
        assert generated.calculation.forceFullCalc
    finally:
        generated.close()
        final.close()


def test_preserves_existing_template_formulas_by_default(tmp_path: Path) -> None:
    output = tmp_path / "consolidated.xlsx"
    config = ConsolidatedPNLWriterConfig(
        cell_writes=[
            CellWrite(
                sheet_name="RAW DATA_Master File",
                cell="B50",
                source_key="raw_master.software_web.total",
            )
        ]
    )

    result = write_consolidated_pnl(
        TEMPLATE,
        output,
        {"raw_master": {"software_web": {"total": 123}}},
        config,
    )

    assert len(result.written_cells) == 0
    assert len(result.skipped_cells) == 1
    assert result.skipped_cells[0]["status"] == "skipped"
    assert result.warnings == ["Skipped formula cell RAW DATA_Master File!B50"]

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["RAW DATA_Master File"]["B50"].value == "=SUM(B46:B49)"
    finally:
        workbook.close()


def test_writes_payroll_art_and_it_visible_formulas_from_allocation_breakdowns(
    tmp_path: Path,
) -> None:
    output = tmp_path / "consolidated.xlsx"
    values = {
        "raw_payroll": {
            "allocation_breakdowns": {
                "art": {
                    "trend_house": 13_007.702,
                    "online": 5_857.694,
                    "dtc": 8_173.08,
                    "general": 5_923.084,
                    "total": 32_961.56,
                },
                "it": {
                    "online": 1_230.768,
                    "dtc": 1_230.768,
                    "general": 13_384.614,
                    "total": 15_846.15,
                },
            }
        }
    }
    config = default_consolidated_pnl_writer_config()
    config.validations = []

    result = write_consolidated_pnl(TEMPLATE, output, values, config)

    assert not result.warnings
    workbook = load_workbook(output, data_only=False)
    try:
        full = workbook["MARCH 2026 FULL "]
        assert full["E44"].value == "=13007.702+F7*5923.084"
        assert full["P44"].value == "=0.0+Q7*5923.084"
        assert full["AA44"].value == "=0.0+AB7*5923.084"
        assert full["AL44"].value == "=5857.694+AM7*5923.084"
        assert full["AW44"].value == "=8173.08+AX7*5923.084"
        assert full["BH44"].value == "=0.0+BI7*5923.084"
        assert full["BV44"].value == "=0.0+BW7*5923.084"

        assert full["E45"].value == "=0.0+F7*13384.614"
        assert full["P45"].value == "=0.0+Q7*13384.614"
        assert full["AA45"].value == "=0.0+AB7*13384.614"
        assert full["AL45"].value == "=1230.768+AM7*13384.614"
        assert full["AW45"].value == "=1230.768+AX7*13384.614"
        assert full["BH45"].value == "=0.0+BI7*13384.614"
        assert full["BV45"].value == "=0.0+BW7*13384.614"
    finally:
        workbook.close()


def test_skips_payroll_art_and_it_formula_writes_without_breakdown_values(
    tmp_path: Path,
) -> None:
    output = tmp_path / "consolidated.xlsx"
    config = default_consolidated_pnl_writer_config()
    config.validations = []

    result = write_consolidated_pnl(
        TEMPLATE,
        output,
        {"raw_payroll": {"sales": {"trend_house": 36_909.05}}},
        config,
    )

    assert not result.warnings
    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["MARCH 2026 FULL "]["E44"].value == "=12526.93+(F7*5884.62)"
        assert any(
            row["sheet_name"] == "MARCH 2026 FULL " and row["cell"] == "E44"
            for row in result.skipped_cells
        )
    finally:
        workbook.close()


def test_validates_written_cells(tmp_path: Path) -> None:
    output = tmp_path / "consolidated.xlsx"
    config = ConsolidatedPNLWriterConfig(
        cell_writes=[
            CellWrite(
                sheet_name="RAW DATA_Payroll",
                cell="B14",
                source_key="raw_payroll.production",
                value_type="number",
                required=True,
            )
        ],
        validations=[
            CellValidation(
                name="production payroll",
                sheet_name="RAW DATA_Payroll",
                cell="B14",
                expected_source_key="raw_payroll.production",
            )
        ],
    )

    result = write_consolidated_pnl(
        TEMPLATE,
        output,
        {"raw_payroll": {"production": "64635.26"}},
        config,
    )

    assert result.written_cells[0]["new_value"] == 64_635.26
    assert result.validation_results[0]["status"] == "ok"
