from pathlib import Path

from openpyxl import Workbook

from hgf_pnl.extractors.pl_by_dept import PLByDeptConfig, extract_pl_by_dept
from hgf_pnl.formulas import FormulaSentinel, WorkbookFormulaEvaluator


def test_evaluates_basic_arithmetic_ranges_and_same_workbook_sheet_refs(tmp_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Main"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=A1+A2*2"
    ws["A4"] = "=SUM(A1:A3)"
    ws["A5"] = "=AVERAGE(A1:A3)"
    other = workbook.create_sheet("Other Sheet")
    other["B1"] = "='Main'!A4+1"

    path = tmp_path / "formulas.xlsx"
    workbook.save(path)

    evaluator = WorkbookFormulaEvaluator.from_path(path)
    try:
        assert evaluator.evaluate_cell("Main", "A3").value == 8
        assert evaluator.evaluate_cell("Main", "A4").value == 13
        assert evaluator.evaluate_cell("Main", "A5").value == 13 / 3
        assert evaluator.evaluate_cell("Other Sheet", "B1").value == 14
    finally:
        evaluator.close()


def test_unsupported_formula_returns_sentinel(tmp_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws["A1"] = 1
    ws["A2"] = '=VLOOKUP(A1,A1:B2,2,FALSE)'

    path = tmp_path / "unsupported.xlsx"
    workbook.save(path)

    evaluator = WorkbookFormulaEvaluator.from_path(path)
    try:
        result = evaluator.evaluate_cell("Sheet", "A2")
        assert result.status == "unsupported"
        assert isinstance(result.value, FormulaSentinel)
        assert result.value.reason == "unsupported_function"
        assert result.value.detail == "VLOOKUP"
    finally:
        evaluator.close()


def test_unsupported_vlookup_with_workbook_index_whole_column_refs_returns_function_sentinel(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws["A1"] = 1
    ws["A2"] = "=VLOOKUP(A1,[1]Sheet1!$A:$B,2,FALSE)"

    path = tmp_path / "unsupported_vlookup_whole_column.xlsx"
    workbook.save(path)

    evaluator = WorkbookFormulaEvaluator.from_path(path)
    try:
        result = evaluator.evaluate_cell("Sheet", "A2")
        assert result.status == "unsupported"
        assert isinstance(result.value, FormulaSentinel)
        assert result.value.reason == "unsupported_function"
        assert result.value.detail == "VLOOKUP"
    finally:
        evaluator.close()


def test_evaluates_workbook_index_sheet_references_as_same_workbook(tmp_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Payroll"
    ws["M56"] = 6176
    ws["M57"] = "=+[1]Payroll!M56"
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "='[1]Payroll'!M56+1"

    path = tmp_path / "workbook_index_refs.xlsx"
    workbook.save(path)

    evaluator = WorkbookFormulaEvaluator.from_path(path)
    try:
        assert evaluator.evaluate_cell("Payroll", "M57").value == 6176
        assert evaluator.evaluate_cell("Summary", "A1").value == 6177
    finally:
        evaluator.close()


def test_pl_by_dept_can_use_calculated_formula_values() -> None:
    sample = Path("sample_files/Workpapers MARCH/DATA/Profit and Loss By Dept.xlsx")
    result = extract_pl_by_dept(
        sample,
        PLByDeptConfig(
            include_total_columns=False,
            calculate_formulas=True,
        ),
    )

    apa_sales = [
        row
        for row in result.rows
        if row["line_item"] == "Sales - APA" and row["department"] == "All Pop Art"
    ]
    assert apa_sales
    assert apa_sales[0]["cached_amount"] == 0
    assert apa_sales[0]["calculated_amount"] == 2786.55
    assert apa_sales[0]["amount"] == 2786.55
    assert apa_sales[0]["calculation_status"] == "ok"
